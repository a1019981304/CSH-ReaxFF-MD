# -*- coding: utf-8 -*-
import os
import re
import math
import csv
from collections import defaultdict, OrderedDict
from statistics import mean, pstdev
import numpy as np
import copy

# 缺失常量定义（保持原代码引用，可根据实际需求调整）
CN_SG_WINDOW = 11 # SG 平滑窗口点数（奇数），不足则降级为移动平均
CN_SG_POLY = 3 # SG 多项式阶数
VALLEY1_RMAX = 4.0 # 第一谷搜索上限（Å）
VALLEY2_RMAX = 7.0 # 第二谷搜索上限（Å）；对 Ca–O 可适当再大些

try:
    # 导入SciPy的Savitzky-Golay滤波（平滑）
    from scipy.signal import savgol_filter
    
    def smooth_sg(y, window=CN_SG_WINDOW, poly=CN_SG_POLY):
        # 确保窗口大小为奇数（Savitzky-Golay滤波要求）
        w = window if window % 2 == 1 else window + 1
        # 窗口大小不超过数据长度（保持奇数特性）
        if len(y) > 2:
            w = min(w, len(y) - (1 - len(y) % 2))
        else:
            w = 3
        # 限制窗口最小为3、多项式阶数最大为5，使用插值模式
        return savgol_filter(y, max(3, w), min(poly, 5), mode='interp')

except Exception:
    # 无SciPy时，退化为移动平均平滑（兼容）
    def smooth_sg(y, window=CN_SG_WINDOW, poly=CN_SG_POLY):
        # 确保窗口大小为奇数
        w = window if window % 2 == 1 else window + 1
        # 窗口大小限制：最小3，最大不超过数据长度的最大奇数
        w = max(3, min(w, len(y) // 2 * 2 + 1))
        # 数据长度不足时直接返回原数据（转换为float类型）
        if w <= 1 or len(y) < w:
            return np.array(y, dtype=float)
        # 边缘填充（避免边界失真）
        pad = w // 2
        ypad = np.pad(y, (pad, pad), mode='edge')
        # 移动平均核函数
        ker = np.ones(w, dtype=float) / w
        # 卷积计算（返回有效区域结果）
        return np.convolve(ypad, ker, mode='valid')

try:
    # 尝试导入SciPy的累积梯形积分函数（效率）
    from scipy.integrate import cumulative_trapezoid as _cumtrapz
    
    def cumtrapz_with_zero(x, y):
        # 累积梯形积分，首元素补0，使输出长度与输入y一致
        return np.insert(_cumtrapz(y, x), 0, 0.0)

except Exception:
    # 无SciPy时，手动实现累积梯形积分（功能一致）
    def cumtrapz_with_zero(x, y):
        # 转换输入为float类型数组
        x = np.asarray(x, dtype=float)
        y = np.asarray(y, dtype=float)
        
        # 输入长度小于2时，返回与x同长度的0数组
        if x.size < 2:
            return np.zeros_like(x, dtype=float)
        
        # 计算相邻x的差值
        dx = np.diff(x)
        # 计算相邻y的平均值
        avg = 0.5 * (y[:-1] + y[1:])
        # 累积求和得到积分结果
        c = np.cumsum(dx * avg)
        # 首元素补0，返回完整积分结果
        return np.concatenate(([0.0], c))

def find_first_two_minima(r, g_smooth, rmax1=VALLEY1_RMAX, rmax2=VALLEY2_RMAX):
    # 只做参考！！！！！
    # 基于导数符号翻转查找前两个极小值（鲁棒、简单实现）
    # 计算梯度（导数近似）
    dr = np.gradient(g_smooth, r)
    
    # 第一步：找到第一个峰值（导数由正转负）
    idx_peak = None
    for i in range(1, len(r) - 1):
        if dr[i-1] > 0 and dr[i] <= 0:
            idx_peak = i
            break
    
    rmin1 = None  # 第一谷值
    rmin2 = None  # 第二谷值
    
    if idx_peak is not None:
        # 第二步：找峰值后的第一个谷值（导数由负转正，且在rmax1范围内）
        for j in range(idx_peak + 1, len(r) - 1):
            if dr[j-1] < 0 and dr[j] >= 0 and r[j] <= rmax1:
                rmin1 = r[j]
                break
        
        # 第三步：继续向后找第二个谷值（导数由负转正，且在rmax2范围内）
        # 起始位置：若未找到第一谷则从峰值后开始，否则从第一谷之后开始
        start_j = idx_peak + 1 if rmin1 is None else np.searchsorted(r, rmin1) + 1
        for j in range(start_j, len(r) - 1):
            if dr[j-1] < 0 and dr[j] >= 0 and r[j] <= rmax2:
                rmin2 = r[j]
                break
    
    return rmin1, rmin2

def get_peak_search_for_pair(pair_name):
    """
    返回 (r_lo_pk, r_hi_pk, dg_min, valley_span)，用于在 g_smooth 上先找第一峰，再在峰后 valley_span 内找第一谷。
    数值按对类别取合理缺省。需要更细粒度时，可以在这里加精确表覆盖。
    """
    # 缺省：类别参数
    # H–O / O–H（氢键）：峰较尖
    H_O = (1.2, 3.0, 0.10, 1.0) # (r_lo, r_hi, Δg_min, 谷搜索宽度)
    # Ca–O：配位/水合
    CA_O = (2.1, 3.2, 0.05, 1.2)
    # Si–O：跨界近邻
    SI_O = (3.2, 4.5, 0.03, 1.2)
    # O–O：氢键网络
    O_O = (3.0, 4.0, 0.05, 1.5)
    # 精确例外（如果需要）
    OVERRIDE = {
        # Ca–O
        "Ca-O_Water_IF_CaCoord": (2.2, 3.0, 0.05, 1.0),
        "Ca-O_SiO2_Structural":  (2.2, 3.0, 0.05, 1.2),
        "Ca-O_SiO2_Surface":     (2.4, 3.3, 0.05, 1.2),
        # H–O 代表性
        "H_SiO2_Hydroxyl-O_Hydroxyl_CSH": (1.3, 2.3, 0.10, 1.0),
        "H_Water_CSH-O_SiO2_Surface":     (1.5, 2.2, 0.10, 1.0),
        "H_Water_IF-O_SiO2_Surface":      (1.5, 2.2, 0.10, 1.0),
        "H_Water_IF-O_Hydroxyl_CSH":      (1.3, 2.2, 0.10, 1.0),
        # Si–O
        "Si_Surface-O_Water_CSH":         (3.4, 4.4, 0.03, 1.4),
    }
    if pair_name in OVERRIDE:
        return OVERRIDE[pair_name]

    # 类型兜底
    left, right = (pair_name.split("-", 1) + [""])[:2]
    def is_H(s):  return s.startswith("H_")
    def is_O(s):  return s.startswith("O_")
    def is_Ca(s): return s.startswith("Ca")
    def is_Si(s): return s.startswith("Si_") or s == "Si"

    if (is_H(left) and is_O(right)) or (is_O(left) and is_H(right)):
        return H_O
    if (is_Ca(left) and is_O(right)) or (is_O(left) and is_Ca(right)):
        return CA_O
    if (is_Si(left) and is_O(right)) or (is_O(left) and is_Si(right)):
        return SI_O
    if (is_O(left) and is_O(right)):
        return O_O
    # 实在不归类
    return (1.2, 6.0, 0.05, 1.2)

def find_first_peak(r, y_smooth, r_lo, r_hi, dg_min):
    """
    在 [r_lo, r_hi] 内找第一个显著峰：导数(+→−) 且振幅超过 dg_min（相对尾部基线）。
    返回 r_peak 或 None。
    """
    mask = (r >= r_lo) & (r <= r_hi)
    if not np.any(mask):
        return None
    rs = r[mask]; ys = y_smooth[mask]
    # 估一个基线：尾部均值（原始窗口高端的后 10%）
    tail_mask = (r >= (r_hi + (r.max()-r_hi)*0.2))
    baseline = float(np.median(y_smooth[tail_mask])) if np.any(tail_mask) else float(np.median(ys))
    dy = np.gradient(ys, rs)
    for i in range(1, len(rs)-1):
        if dy[i-1] > 0 and dy[i] <= 0:
            if (ys[i] - baseline) >= dg_min:
                return float(rs[i])
    return None

def find_min_after_peak(r, y_smooth, r_peak, span):
    """
    在 (r_peak, r_peak+span] 内找 y_smooth 的最小值位置。
    """
    if r_peak is None:
        return None
    lo = r_peak + 1e-6
    hi = r_peak + span
    mask = (r >= lo) & (r <= hi)
    if np.count_nonzero(mask) < 3:
        return None
    sub_r = r[mask]; sub_y = y_smooth[mask]
    j = int(np.argmin(sub_y))
    return float(sub_r[j])

def get_valley_windows_for_pair(pair_name,
                               default1=(VALLEY1_RMAX*0.5, VALLEY1_RMAX),
                               default2=(VALLEY1_RMAX, VALLEY2_RMAX)):
    """
    返回 ((lo1, hi1), (lo2, hi2)) in Å：在 g_smooth 上分别在两个窗口内找第一/第二谷。
    先查精确表 → 再按类型规则 → 最后用兜底默认。
    """
    # 精确表
    PAIR_WINDOWS = {
        # Ca–O（第一谷紧跟首峰 ~2.8–3.3）
        "Ca-O_SiO2_Structural": ((2.80, 3.60), (3.60, 6.50)),
        "Ca-O_SiO2_Surface": ((2.80, 3.60), (3.60, 6.50)),
        "Ca-O_Water_IF_CaCoord": ((2.60, 3.30), (3.30, 6.50)),
        # H–O（氢键，也覆盖 O–H）
        "H_SiO2_Hydroxyl-O_Hydroxyl_CSH": ((2.20, 2.80), (2.80, 4.00)),
        "H_SiO2_Hydroxyl-O_Water_CSH":    ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_IF-O_SiO2_Surface":      ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_IF-O_SiO2_Structural":   ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_IF-O_Hydroxyl_CSH":      ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_IF-O_NBO_CSH":           ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_CSH-O_SiO2_Surface":     ((2.20, 2.80), (2.80, 4.00)),
        "H_Water_CSH-O_SiO2_Structural":  ((2.20, 2.80), (2.80, 4.00)),
        "O_Water_IF-H_SiO2_Hydroxyl":     ((2.20, 2.80), (2.80, 4.00)),
        "O_Water_IF-H_Hydroxyl_CSH":      ((2.20, 2.80), (2.80, 4.00)),
        # O–O（氢键网络）
        "O_SiO2_Surface-O_Hydroxyl_CSH":  ((3.20, 3.80), (3.80, 6.00)),
        "O_Water_IF-O_SiO2_Surface":      ((3.20, 3.80), (3.80, 6.00)),
        "O_Water_IF-O_Hydroxyl_CSH":      ((3.20, 3.80), (3.80, 6.00)),
        # Si–O（跨界近邻）
        "Si_Surface-O_Water_CSH":         ((3.40, 4.20), (4.20, 7.00)),
        "Si_Bulk-O_Water_CSH":            ((3.40, 4.20), (4.20, 7.00)),
        # 汇总/混合集
        "H_SiO2_Hydroxyl-O_CSH_all":      ((2.20, 2.80), (2.80, 4.00)),
    }
    if pair_name in PAIR_WINDOWS:
        return PAIR_WINDOWS[pair_name]

    # 类型规则兜底
    if "-" in pair_name:
        left, right = pair_name.split("-", 1)
    else:
        left, right = pair_name, ""
    
    def is_H(s):  return s.startswith("H_")
    def is_O(s):  return s.startswith("O_")
    def is_Ca(s): return s.startswith("Ca")
    def is_Si(s): return s.startswith("Si_") or s == "Si"

    # H–O / O–H
    if (is_H(left) and is_O(right)) or (is_O(left) and is_H(right)):
        return ((2.20, 2.80), (2.80, 4.00))
    # O–O
    if is_O(left) and is_O(right):
        return ((3.20, 3.80), (3.80, 6.00))
    # Ca–O
    if (is_Ca(left) and is_O(right)) or (is_O(left) and is_Ca(right)):
        return ((2.80, 3.60), (3.60, 6.50))
    # Si–O
    if (is_Si(left) and is_O(right)) or (is_O(left) and is_Si(right)):
        return ((3.40, 4.20), (4.20, 7.00))

    # 全局默认
    return (default1, default2)



import math
import numpy as np
import csv
import os

# smooth_sg、cumtrapz_with_zero、get_peak_search_for_pair、find_first_peak、find_min_after_peak、get_valley_windows_for_pair、fnum

def compute_rdf_cn_for_all_pairs(rdf_results, nb_block_means_by_pair, veff_block_means, out_dir, pair_tag_map=None):
    # 1) ρB：每块 NB_block_mean / Veff_block_mean → 块均值±SE
    n_blocks = len(veff_block_means)
    rhoB_stats = {}
    for pair_name, res in rdf_results.items():
        if pair_name not in nb_block_means_by_pair or n_blocks == 0:
            continue
        nb_blocks = nb_block_means_by_pair[pair_name]
        m = min(len(nb_blocks), n_blocks)
        if m == 0:
            continue
        rhos = np.array([nb_blocks[b] / veff_block_means[b] for b in range(m)], dtype=float)
        rho_mean = float(np.mean(rhos))
        rho_se = float(np.std(rhos, ddof=1) / math.sqrt(m)) if m > 1 else 0.0
        rhoB_stats[pair_name] = (rho_mean, rho_se, m)

    # 2) CN(r)：CN_raw 用 g_mean；CN_smooth 用 g_smooth。
    summary_rows = []
    for pair_name, res in rdf_results.items():
        if pair_name not in rhoB_stats:
            continue
        rho_mean, rho_se, m = rhoB_stats[pair_name]
        r   = np.asarray(res["r"], dtype=float)
        g   = np.asarray(res["g_mean"], dtype=float)
        gse = np.asarray(res["g_se"],   dtype=float)

        # 平滑
        g_smooth = smooth_sg(g, CN_SG_WINDOW, CN_SG_POLY)

        # CN 曲线（用 g_mean 积分）
        cn_raw    = 4.0 * math.pi * rho_mean * cumtrapz_with_zero(r, g        * (r**2))
        cn_raw_se = 4.0 * math.pi * rho_mean * cumtrapz_with_zero(r, gse      * (r**2))
        cn_smooth = 4.0 * math.pi * rho_mean * cumtrapz_with_zero(r, g_smooth * (r**2))

        # 先找第一峰，峰后找第一谷
        r_lo_pk, r_hi_pk, dg_min, vspan = get_peak_search_for_pair(pair_name)
        r_pk   = find_first_peak(r, g_smooth, r_lo_pk, r_hi_pk, dg_min)
        rmin1  = find_min_after_peak(r, g_smooth, r_pk, vspan)

        # 兜底：若先峰后谷失败，回退到窗口法（可选）
        if rmin1 is None:
            try:
                (win1_lo, win1_hi), _ = get_valley_windows_for_pair(pair_name)
                mask = (r >= win1_lo) & (r <= win1_hi)
                if np.count_nonzero(mask) >= 3:
                    sub_r = r[mask]; sub_y = g_smooth[mask]
                    rmin1 = float(sub_r[int(np.argmin(sub_y))])
            except Exception:
                rmin1 = None

        # 可选：找第二谷（用于报告 second-shell）
        rmin2 = None
        if rmin1 is not None:
            rmin2 = find_min_after_peak(r, g_smooth, rmin1, vspan*2.0)

        def cn_at(r_cut, cn_curve):
            if r_cut is None: return None
            i = int(np.argmin(np.abs(r - r_cut)))
            return float(cn_curve[i])

        cn1_raw    = cn_at(rmin1, cn_raw)
        cn1_smooth = cn_at(rmin1, cn_smooth)
        cn2_raw    = cn_at(rmin2, cn_raw)
        cn2_smooth = cn_at(rmin2, cn_smooth)

        # 3) 每对导出：r, g_mean, g_se, CN_raw, CN_raw_se, g_smooth, CN_smooth
        tag = pair_tag_map.get(pair_name, "") if pair_tag_map else ""
        prefix = f"rdf_cn_{(tag + '_') if tag else ''}{pair_name}"
        csv_path = os.path.join(out_dir, f"{prefix}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, lineterminator="\r\n")
            w.writerow(["r(Angstrom)", "g_mean", "g_se",
                        "CN_raw", "CN_raw_se", "g_smooth", "CN_smooth"])
            for i in range(len(r)):
                w.writerow([fnum(r[i]), fnum(g[i]), fnum(gse[i]),
                            fnum(cn_raw[i]), fnum(cn_raw_se[i]),
                            fnum(g_smooth[i]), fnum(cn_smooth[i])])

        # 汇总
        summary_rows.append([
            pair_name, fnum(rho_mean), fnum(rho_se), int(m),
            fnum(rmin1) if rmin1 is not None else "",
            fnum(cn1_raw) if cn1_raw is not None else "",
            fnum(cn1_smooth) if cn1_smooth is not None else "",
            fnum(rmin2) if rmin2 is not None else "",
            fnum(cn2_raw) if cn2_raw is not None else "",
            fnum(cn2_smooth) if cn2_smooth is not None else "",
        ])

    # 4) 汇总表
    if summary_rows:
        summary_path = os.path.join(out_dir, "rdf_cn_summary.csv")
        with open(summary_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, lineterminator="\r\n")
            w.writerow(["pair", "rho_B_mean(1/Å^3)", "rho_B_se", "n_blocks",
                        "r_min1(Å)", "CN_raw@min1", "CN_smooth@min1",
                        "r_min2(Å)", "CN_raw@min2", "CN_smooth@min2"])
            for row in summary_rows:
                w.writerow(row)
# =============================================================================
# 全局配置
# =============================================================================
WORK_PATH = r"F:\MD\python-code\水键转换\H2O--5A"
DATA_PATTERN = os.path.join(WORK_PATH, "data_frame_classified_{i}.data")

# 帧控制
N_FRAMES = 200          # 200帧
START_FRAME = 0
FRAME_STEP = 1

# 块平均
BLOCK_SIZE = 10         

# 氢键几何阈值
HBOND_R_OH_MAX = 2.45  # Å
HBOND_ANGLE_MIN = 150.0  # degree
NEAR_O_FALLBACK = 1.25    # 找母氧的几何兜底阈值（Å）

# 氢键判据开关
HBOND_USE_ANGLE   = True     # False 时不判角度，只用 H···O 距离
HBOND_USE_OO_CUT  = False    # True 时同时要求 O_d···O_a ≤ HBOND_OO_MAX
HBOND_OO_MAX      = 3.30     # Å，常见几何阈值 3.2~3.5

##稀有键诊断开关
ENABLE_DIAG_SI_TO_CSH_WATER = False   # Si_*–O_Bridging_CSH / O_Water_CSH（恒0或极少）
ENABLE_CA_O_WATER_CSH       = False   # Ca–O_Water_CSH（如关心 CSH 水配位可打开）

# RDF 参数
RDF_DR = 0.02  # 或 0.05

#界面定位分位数
IF_Q_CSH_TOP = 0.95
IF_Q_SIO2_BOT = 0.05

#界面窗口（总宽度的一半，Å）
IF_WINDOW_HALF = 6.0

#是否对化学键/共价 OH 也应用界面窗口
APPLY_WINDOW_TO_BONDS = True # True/False

#RDF 范围到 10 Å
RDF_R_MAX = 10.0 # 覆盖原值8.0

#one-per-donor 选择策略的兜底 tie-break：按原子ID
TIE_BREAK_USE_ID = True

# 类型常量（与 data 文件一致）
Ca = 1
CSH_Si = 2
O_Bridging_CSH = 3
O_Water_CSH = 4
O_Hydroxyl_CSH = 5
O_NBO_CSH = 6
O_Unclassified_CSH = 7
O_SiO2_Structural = 8
O_SiO2_Surface = 9
Si_Bulk = 10
Si_Surface = 11
H_Water_CSH = 12
H_Hydroxyl_CSH = 13
H_SiO2_Hydroxyl = 14
H_Unbonded_CSH = 15
H_Unbonded_SiO2 = 16
O_Water_IF = 17
O_Water_IF_CaCoord = 18
H_Water_IF = 19

TYPE_NAME = {
    Ca: "Ca",
    CSH_Si: "CSH_Si",
    O_Bridging_CSH: "O_Bridging_CSH",
    O_Water_CSH: "O_Water_CSH",
    O_Hydroxyl_CSH: "O_Hydroxyl_CSH",
    O_NBO_CSH: "O_NBO_CSH",
    O_Unclassified_CSH: "O_Unclassified_CSH",
    O_SiO2_Structural: "O_SiO2_Structural",
    O_SiO2_Surface: "O_SiO2_Surface",
    Si_Bulk: "Si_Bulk",
    Si_Surface: "Si_Surface",
    H_Water_CSH: "H_Water_CSH",
    H_Hydroxyl_CSH: "H_Hydroxyl_CSH",
    H_SiO2_Hydroxyl: "H_SiO2_Hydroxyl",
    H_Unbonded_CSH: "H_Unbonded_CSH",
    H_Unbonded_SiO2: "H_Unbonded_SiO2",
    O_Water_IF: "O_Water_IF",
    O_Water_IF_CaCoord: "O_Water_IF_CaCoord",
    H_Water_IF: "H_Water_IF",
}

CA_O_KEYS = ["Ca-O_SiO2_Surface", "Ca-O_SiO2_Structural", "Ca-O_Water_IF_CaCoord"]

AMU_TO_G = 1.66053906660e-24 # g/amu
ANG3_TO_CM3 = 1.0e-24 # cm^3/Å^3

TOTAL_MASS = 0
TYPE_NAME[TOTAL_MASS] = "Total_mass"

MASS_AMU = {
    Ca: 40.079800,CSH_Si: 28.085500, Si_Bulk: 28.085500, Si_Surface: 28.085500,

    O_Bridging_CSH: 15.999400, O_Water_CSH: 15.999400, O_Hydroxyl_CSH: 15.999400,
    O_NBO_CSH: 15.999400, O_Unclassified_CSH: 15.999400,
    O_SiO2_Structural: 15.999400, O_SiO2_Surface: 15.999400,
    # 有水脚本多出来的两类
    O_Water_IF: 15.999400, O_Water_IF_CaCoord: 15.999400,

    H_Water_CSH: 1.007970, H_Hydroxyl_CSH: 1.007970, H_SiO2_Hydroxyl: 1.007970,
    H_Unbonded_CSH: 1.007970, H_Unbonded_SiO2: 1.007970,
    # 有水脚本的 H_Water_IF
    H_Water_IF: 1.007970,
}

# 集合
HYDROGEN_TYPES = {H_Water_CSH, H_Hydroxyl_CSH, H_SiO2_Hydroxyl, H_Unbonded_CSH, H_Unbonded_SiO2, H_Water_IF}
OXYGEN_TYPES = {O_Bridging_CSH, O_Water_CSH, O_Hydroxyl_CSH, O_NBO_CSH, O_Unclassified_CSH, O_SiO2_Structural, O_SiO2_Surface, O_Water_IF, O_Water_IF_CaCoord}
SILICON_TYPES = {CSH_Si, Si_Bulk, Si_Surface}

# =============================================================================
# 工具与几何
# =============================================================================
def parse_lammps_data(path):
    """
    读取 LAMMPS data 文件：box_bounds, atoms, bonds
    Atoms 行格式（与你分类写出的 data 一致）:
      id 1 type charge x y z ix iy iz
    Bonds 行格式：
      bid 1 i j
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"data file not found: {path}")
    with open(path, 'r') as f:
        lines = f.readlines()

    # 解析 header 中的 atom/bond 数与 box_bounds
    n_atoms = None
    n_bonds = 0
    box_bounds = []
    i = 0

    # 找到 counts
    while i < len(lines):
        line = lines[i].strip()
        if re.search(r'^\d+\s+atoms', line):
            n_atoms = int(line.split()[0])
        elif re.search(r'^\d+\s+bonds', line):
            n_bonds = int(line.split()[0])
        elif "xlo xhi" in line:
            # box bounds 
            xlo, xhi = map(float, lines[i].split()[:2])
            ylo, yhi = map(float, lines[i+1].split()[:2])
            zlo, zhi = map(float, lines[i+2].split()[:2])
            box_bounds = [(xlo, xhi), (ylo, yhi), (zlo, zhi)]
            i += 2
        i += 1

    if n_atoms is None or len(box_bounds) != 3:
        raise ValueError("Failed to parse header/box bounds/atom count from data file.")

    def find_line_idx(keyword):
        for idx, ln in enumerate(lines):
            if ln.strip().startswith(keyword):
                return idx
        return -1

    atoms_idx = find_line_idx("Atoms")
    if atoms_idx == -1:
        raise ValueError("Atoms section not found.")

    # 读取 Atoms
    atoms = []
    i = atoms_idx + 2  # 跳过 "Atoms" 和一个空行
    read_atoms = 0
    while i < len(lines) and read_atoms < n_atoms:
        parts = lines[i].split()
        if len(parts) >= 10:
            atom_id = int(parts[0])
            atype = int(parts[2])
            # charge = float(parts[3])
            x = float(parts[4]); y = float(parts[5]); z = float(parts[6])
            # ix,iy,iz = int(parts[7]), int(parts[8]), int(parts[9])
            atoms.append((atom_id, atype, (x, y, z)))
            read_atoms += 1
        i += 1

    # 读取 Bonds
    bonds = []
    bonds_idx = find_line_idx("Bonds")
    if bonds_idx != -1 and n_bonds > 0:
        j = bonds_idx + 2
        read_bonds = 0
        while j < len(lines) and read_bonds < n_bonds:
            parts = lines[j].split()
            if len(parts) >= 4:
                a1 = int(parts[2]); a2 = int(parts[3])
                if a1 != a2:
                    # 无向去重留给上层
                    bonds.append((min(a1,a2), max(a1,a2)))
                read_bonds += 1
            j += 1

    # 生成索引
    pos = {}
    typ = {}
    ids_by_type = defaultdict(list)
    for aid, atype, xyz in atoms:
        pos[aid] = xyz
        typ[aid] = atype
        ids_by_type[atype].append(aid)

    bonds_set = set(bonds)
    adj = defaultdict(set)
    for a, b in bonds_set:
        adj[a].add(b)
        adj[b].add(a)

    return {
        'box': tuple(box_bounds),    # ((xlo,xhi), (ylo,yhi), (zlo,zhi))
        'pos': pos,                  # dict id -> (x,y,z)
        'typ': typ,                  # dict id -> type
        'ids_by_type': ids_by_type,  # dict type -> [ids]
        'bonds': bonds_set,          # set of (i,j)
        'adj': adj                   # dict id -> set(ids)
    }

def box_lengths(box):
    (xlo, xhi), (ylo, yhi), (zlo, zhi) = box
    return (xhi - xlo, yhi - ylo, zhi - zlo)

def min_image_delta(dx, L):
    if L <= 0: 
        return dx
    dx -= L * round(dx / L)
    return dx

def dist_pbc(p1, p2, box):
    Lx, Ly, Lz = box_lengths(box)
    dx = min_image_delta(p1[0] - p2[0], Lx)
    dy = min_image_delta(p1[1] - p2[1], Ly)
    dz = min_image_delta(p1[2] - p2[2], Lz)
    return math.sqrt(dx*dx + dy*dy + dz*dz)


def angle_Od_H_Oa(Od, H, Oa, box):
    # ∠(Od–H···Oa)，两条射线都从 H 出发（H 为顶点）
    Lx, Ly, Lz = box_lengths(box)
    # 正确方向：u = Od - H；v = Oa - H
    u = (min_image_delta(Od[0]-H[0], Lx), min_image_delta(Od[1]-H[1], Ly), min_image_delta(Od[2]-H[2], Lz))
    v = (min_image_delta(Oa[0]-H[0], Lx), min_image_delta(Oa[1]-H[1], Ly), min_image_delta(Oa[2]-H[2], Lz))
    nu = math.sqrt(u[0]**2 + u[1]**2 + u[2]**2) + 1e-12
    nv = math.sqrt(v[0]**2 + v[1]**2 + v[2]**2) + 1e-12
    dot = (u[0]*v[0] + u[1]*v[1] + u[2]*v[2]) / (nu * nv)
    dot = max(-1.0, min(1.0, dot))
    return math.degrees(math.acos(dot))

def bin_index(r, rmax, dr):
    if r < 0 or r >= rmax:
        return -1
    return int(r // dr)

def fnum(x, nd=10):

    try:
        # 将输入转换为浮点数后，按指定有效数字位数格式化
        return f"{float(x):.{nd}g}"
    except Exception:
        # 转换失败时，直接返回原始值的字符串形式
        return str(x)
    

def compute_interface_markers(frame_data, q_csh=IF_Q_CSH_TOP, q_sio2=IF_Q_SIO2_BOT):
    pos = frame_data['pos']; ids_by_type = frame_data['ids_by_type']
    # CSH 顶面：CSH_Si (Type 2)
    z_csh = [pos[i][2] for i in ids_by_type.get(CSH_Si, [])]
    # SiO2 底面：所有 Si（Si_Bulk + Si_Surface）
    z_si = []
    for t in (Si_Bulk, Si_Surface):
        z_si.extend([pos[i][2] for i in ids_by_type.get(t, [])])
    if not z_csh or not z_si:
        raise ValueError("interface markers: not enough atoms in CSH_Si or Si (bulk/surface)")
    z_csh_top = float(np.quantile(np.asarray(z_csh), q_csh))
    z_sio2_bot = float(np.quantile(np.asarray(z_si), q_sio2))
    z0 = 0.5 * (z_csh_top + z_sio2_bot)
    deq = z_sio2_bot - z_csh_top # 弛豫后的间距（正值）
    return z_csh_top, z_sio2_bot, z0, deq


def make_aligned_frame(frame_data, z0):
    fd = {
    'box': frame_data['box'],
    'pos': {},
    'typ': frame_data['typ'],
    'ids_by_type': frame_data['ids_by_type'],
    'bonds': frame_data['bonds'],
    'adj': frame_data['adj'],
    }
    for aid, (x, y, z) in frame_data['pos'].items():
        fd['pos'][aid] = (x, y, z - z0)
    (xlo, xhi), (ylo, yhi), (zlo, zhi) = frame_data['box']
    fd['box'] = ((xlo, xhi), (ylo, yhi), (zlo - z0, zhi - z0))
    return fd

def window_mask_ids(frame_data_aligned, half_w=IF_WINDOW_HALF):
    pos = frame_data_aligned['pos']
    mask = set(aid for aid, p in pos.items() if abs(p[2]) <= half_w)
    return mask

def filter_frame_for_window(frame_data_aligned, mask_ids, filter_bonds=False):
    ids_by_type_new = defaultdict(list)
    for t, lst in frame_data_aligned['ids_by_type'].items():
        ids_by_type_new[t] = [aid for aid in lst if aid in mask_ids]
    bonds_new = frame_data_aligned['bonds']
    if filter_bonds:
        bonds_new = set((i, j) for (i, j) in frame_data_aligned['bonds'] if i in mask_ids and j in mask_ids)
    return {
        'box': frame_data_aligned['box'],
        'pos': frame_data_aligned['pos'],
        'typ': frame_data_aligned['typ'],
        'ids_by_type': ids_by_type_new,
        'bonds': bonds_new,
        'adj': frame_data_aligned['adj'],
}

def area_xy_nm2(frame_data):
    (xlo,xhi),(ylo,yhi),_ = frame_data['box']
    A_xy_ang2 = (xhi - xlo) * (yhi - ylo)
    return A_xy_ang2 * 1.0e-2 # Å^2 -> nm^2
# =============================================================================
# 分析：化学键计数（来自 Bonds；不含 O–H）
# =============================================================================
def count_chemical_bonds(frame_data):
    """
    只统计“化学键”里的 Si–O、Ca–O（不含任何 H 相关对）
    主表聚焦“界面相关且物理上合理”的键；
    另外提供两个诊断开关，可按需展开“几乎为0/稀有事件”的额外对。
    """
    typ = frame_data['typ']; bonds = frame_data['bonds']
    # 诊断开关（需要“看见 0”或排查稀有事件时再打开）
    #ENABLE_DIAG_SI_TO_CSH_WATER = False   # Si_*–O_Bridging_CSH / O_Water_CSH（恒0或极少）
    #ENABLE_CA_O_WATER_CSH       = False   # Ca–O_Water_CSH（如关心 CSH 水配位可打开）
    # 定义统计对（用无序集合匹配）
    PAIRS = OrderedDict({

        # Ca 与 SiO2
        "Ca-O_SiO2_Surface":        ({Ca, O_SiO2_Surface},),
        "Ca-O_SiO2_Structural":     ({Ca, O_SiO2_Structural},),
        # IF 水（被 Ca 配位）
        "Ca-O_Water_IF_CaCoord":    ({Ca, O_Water_IF_CaCoord},),

        # 跨界面 Si–O（CSH 框架 Si 与 SiO2 氧）
        "CSH_Si-O_SiO2_Surface":    ({CSH_Si, O_SiO2_Surface},),
        "CSH_Si-O_SiO2_Structural": ({CSH_Si, O_SiO2_Structural},),

        # SiO2 侧 Si 与 CSH 活性氧（重构/接枝等）
        "Si_Surface-O_Hydroxyl_CSH":({Si_Surface, O_Hydroxyl_CSH},),
        "Si_Bulk-O_Hydroxyl_CSH":   ({Si_Bulk,    O_Hydroxyl_CSH},),
        "Si_Surface-O_NBO_CSH":     ({Si_Surface, O_NBO_CSH},),
        "Si_Bulk-O_NBO_CSH":        ({Si_Bulk,    O_NBO_CSH},),
    })
    # 氢类型：O–H不计入化学键
    # 可选：Ca–O_Water_CSH（如需观察 CSH 内部水对 Ca 的配位）
    if ENABLE_CA_O_WATER_CSH:
        PAIRS.update(OrderedDict({
            "Ca-O_Water_CSH": ({Ca, O_Water_CSH},),
        }))

    # 可选诊断：Si_* 与 CSH 桥氧/水氧的“化学键”（通常应为 0，异常反应用）
    if ENABLE_DIAG_SI_TO_CSH_WATER:
        PAIRS.update(OrderedDict({
            "Si_Surface-O_Bridging_CSH": ({Si_Surface, O_Bridging_CSH},),
            "Si_Bulk-O_Bridging_CSH":    ({Si_Bulk,    O_Bridging_CSH},),
            "Si_Surface-O_Water_CSH":    ({Si_Surface, O_Water_CSH},),
            "Si_Bulk-O_Water_CSH":       ({Si_Bulk,    O_Water_CSH},),
        }))
    counts = {k: 0 for k in PAIRS.keys()}

    for (i, j) in bonds:
        t1 = typ.get(i, -1); t2 = typ.get(j, -1)
        if t1 in HYDROGEN_TYPES or t2 in HYDROGEN_TYPES:
            continue
        s = {t1, t2}
        for name, sets in PAIRS.items():
            for ss in sets:
                if s == ss:
                    counts[name] += 1
                    break
    return counts

# =============================================================================
# 分析：O–H 共价键计数（来自 Bonds）
# =============================================================================
def count_oh_bonds(frame_data):
    
    typ = frame_data['typ']; bonds = frame_data['bonds']

    OH_PAIRS = OrderedDict({
        "OH_SiO2": ({H_SiO2_Hydroxyl, O_SiO2_Surface},),
        "OH_CSH_on_O_Hydroxyl": ({H_Hydroxyl_CSH, O_Hydroxyl_CSH},),
        "OH_CSH_on_O_Bridging": ({H_Hydroxyl_CSH, O_Bridging_CSH},),
        "OH_CSH_Water": ({H_Water_CSH, O_Water_CSH},),
        "OH_IF": ({H_Water_IF, O_Water_IF},),
        "OH_IF_CaCoord": ({H_Water_IF, O_Water_IF_CaCoord},),
    })
    counts = {k: 0 for k in OH_PAIRS.keys()}

    for (i, j) in bonds:
        t1 = typ.get(i, -1); t2 = typ.get(j, -1)
        s = {t1, t2}
        for name, sets in OH_PAIRS.items():
            for ss in sets:
                if s == ss:
                    counts[name] += 1
                    break
    return counts

# =============================================================================
# 分析：氢键计数（几何）
# =============================================================================
def count_hbonds(frame_data):
    """
    返回三套计数：
      - dist_only: 只看 H···O ≤ HBOND_R_OH_MAX
      - dist_oo:   同时要求 O_d···O_a ≤ HBOND_OO_MAX（当 HBOND_USE_OO_CUT=True 时有效，否则与 dist_only 相同）
      - full:      在 dist_oo 基础上再加角度（当 HBOND_USE_ANGLE=True 时有效，否则与 dist_oo 相同）
    """
    box = frame_data['box']; pos = frame_data['pos']; typ = frame_data['typ']
    ids_by_type = frame_data['ids_by_type']; adj = frame_data['adj']

    donor_H_types = {H_SiO2_Hydroxyl, H_Hydroxyl_CSH, H_Water_IF, H_Water_CSH}
    acceptor_O_types = {O_SiO2_Surface, O_SiO2_Structural, O_Hydroxyl_CSH, O_NBO_CSH,
                        O_Water_CSH, O_Water_IF, O_Water_IF_CaCoord, O_Bridging_CSH}


    PAIRS = OrderedDict({

        # SiO2 表面羟基供体 → CSH 侧（含 CSH 水）与 IF 水
        "H_SiO2_Hydroxyl->O_Bridging_CSH": (H_SiO2_Hydroxyl, O_Bridging_CSH),
        "H_SiO2_Hydroxyl->O_Water_CSH": (H_SiO2_Hydroxyl, O_Water_CSH),
        "H_SiO2_Hydroxyl->O_Hydroxyl_CSH": (H_SiO2_Hydroxyl, O_Hydroxyl_CSH),
        "H_SiO2_Hydroxyl->O_NBO_CSH": (H_SiO2_Hydroxyl, O_NBO_CSH),
        "H_SiO2_Hydroxyl->O_Water_IF": (H_SiO2_Hydroxyl, O_Water_IF),
        "H_SiO2_Hydroxyl->O_Water_IF_CaCoord":(H_SiO2_Hydroxyl, O_Water_IF_CaCoord),

        # CSH 水/羟基供体 → SiO2 侧 + IF 水
        "H_Water_CSH->O_SiO2_Structural":     (H_Water_CSH, O_SiO2_Structural),
        "H_Water_CSH->O_SiO2_Surface":        (H_Water_CSH, O_SiO2_Surface),
        "H_Water_CSH->O_Water_IF":            (H_Water_CSH, O_Water_IF),
        "H_Water_CSH->O_Water_IF_CaCoord":    (H_Water_CSH, O_Water_IF_CaCoord),

        "H_Hydroxyl_CSH->O_SiO2_Structural":  (H_Hydroxyl_CSH, O_SiO2_Structural),
        "H_Hydroxyl_CSH->O_SiO2_Surface":     (H_Hydroxyl_CSH, O_SiO2_Surface),
        "H_Hydroxyl_CSH->O_Water_IF":         (H_Hydroxyl_CSH, O_Water_IF),
        "H_Hydroxyl_CSH->O_Water_IF_CaCoord": (H_Hydroxyl_CSH, O_Water_IF_CaCoord),

        # IF 水供体 → 两侧受体
        "H_Water_IF->O_SiO2_Structural":      (H_Water_IF, O_SiO2_Structural),
        "H_Water_IF->O_SiO2_Surface":         (H_Water_IF, O_SiO2_Surface),
        "H_Water_IF->O_Bridging_CSH":         (H_Water_IF, O_Bridging_CSH),
        "H_Water_IF->O_Water_CSH":            (H_Water_IF, O_Water_CSH),
        "H_Water_IF->O_Hydroxyl_CSH":         (H_Water_IF, O_Hydroxyl_CSH),
        "H_Water_IF->O_NBO_CSH":              (H_Water_IF, O_NBO_CSH),

        # 可选：IF 水内水-水
        "H_Water_IF->O_Water_IF":             (H_Water_IF, O_Water_IF),
        "H_Water_IF->O_Water_IF_CaCoord":     (H_Water_IF, O_Water_IF_CaCoord),

      
    })

    # 三套计数
    counts_dist_only = {k: 0 for k in PAIRS.keys()}
    counts_dist_oo   = {k: 0 for k in PAIRS.keys()}
    counts_full      = {k: 0 for k in PAIRS.keys()}

    # donor IDs
    donors_by_type = {}
    for ht in donor_H_types:
        donors_by_type[ht] = ids_by_type.get(ht, [])

    # 受体 IDs
    acceptors_by_type = {}
    for ot in acceptor_O_types:
        acceptors_by_type[ot] = ids_by_type.get(ot, [])

    # 处理每个 donor H
    for ht, donor_ids in donors_by_type.items():
        for hid in donor_ids:
            hpos = pos[hid]

            # 找母氧 O_d（先 bonds 再兜底）
            Od = None
            for nb in adj.get(hid, ()):
                if typ.get(nb) in OXYGEN_TYPES:
                    Od = nb; break
            if Od is None:
                best_r = 1e9; best_id = None
                for ot in OXYGEN_TYPES:
                    for oid in ids_by_type.get(ot, ()):
                        r = dist_pbc(hpos, pos[oid], box)
                        if r < best_r:
                            best_r = r; best_id = oid
                if best_id is None or best_r > NEAR_O_FALLBACK:
                    continue
                Od = best_id
            Od_pos = pos[Od]

            # 对每个定义的 pair，遍历相应受体
            for name, (H_need, O_need) in PAIRS.items():
                if ht != H_need:
                    continue
                for Oa in acceptors_by_type.get(O_need, []):
                    if Oa == Od:
                        continue
                    Oa_pos = pos[Oa]

                    # 1) 只判 H···O
                    r_HA = dist_pbc(hpos, Oa_pos, box)
                    if r_HA > HBOND_R_OH_MAX:
                        continue
                    counts_dist_only[name] += 1

                    # 2) 判 O_d···O_a（若开关打开，否则等于 dist_only）
                    passed_oo = True
                    if HBOND_USE_OO_CUT:
                        r_OO = dist_pbc(Od_pos, Oa_pos, box)
                        if r_OO > HBOND_OO_MAX:
                            passed_oo = False
                    if passed_oo:
                        counts_dist_oo[name] += 1
                    else:
                        continue  # OO 不通过就不再判角度

                    # 3) 判角度（若开关打开，否则等于 dist_oo）
                    if HBOND_USE_ANGLE:
                        ang = angle_Od_H_Oa(Od_pos, hpos, Oa_pos, box)
                        if ang < HBOND_ANGLE_MIN:
                            continue
                    counts_full[name] += 1


    frame_counts = counts_full

    # 可选：在第0帧额外打印一次分级计数（不破坏CSV）
    return frame_counts, counts_dist_only, counts_dist_oo

def count_hbonds_one_per_donor_window(frame_data_aligned, mask_ids):
    #"""
    #在窗口内统计 one-per-donor 氢键：
    #- donor H 与 acceptor O 都需在窗口内
    #- 先过几何阈值（H···O，OO可选，角度可选），在候选中选“最佳受体”：
    #角度越大越优先；再以 H···O 更近；再以 O···O 更近；最后按ID（可控）打破平手
    #- 返回：full_counts（按 PAIRS 名称），并返回归一化/面积（#/nm^2）
    #"""
    box = frame_data_aligned['box']; pos = frame_data_aligned['pos']; typ = frame_data_aligned['typ']
    ids_by_type = frame_data_aligned['ids_by_type']; adj = frame_data_aligned['adj']

    donor_H_types = {H_SiO2_Hydroxyl, H_Hydroxyl_CSH, H_Water_IF, H_Water_CSH}
    acceptor_O_types = {O_SiO2_Surface, O_SiO2_Structural, O_Hydroxyl_CSH, O_NBO_CSH,
                        O_Water_CSH, O_Water_IF, O_Water_IF_CaCoord, O_Bridging_CSH}

    #只记跨界面互作常见项
 
    PAIRS = OrderedDict({
        # --- 1. 界面水 (H_Water_IF) 向下抓 SiO2 ---
        "H_Water_IF->O_SiO2_Surface":        (H_Water_IF, O_SiO2_Surface),
        "H_Water_IF->O_SiO2_Structural":     (H_Water_IF, O_SiO2_Structural),
        
        # --- 2. 界面水 (H_Water_IF) 向上抓 CSH  ---
        "H_Water_IF->O_Hydroxyl_CSH":        (H_Water_IF, O_Hydroxyl_CSH),
        "H_Water_IF->O_NBO_CSH":             (H_Water_IF, O_NBO_CSH),
        "H_Water_IF->O_Bridging_CSH":        (H_Water_IF, O_Bridging_CSH),  
        "H_Water_IF->O_Water_CSH":           (H_Water_IF, O_Water_CSH),     
        
        # --- 3. 界面水 (H_Water_IF) 内部互作  ---
        "H_Water_IF->O_Water_IF":            (H_Water_IF, O_Water_IF),
        "H_Water_IF->O_Water_IF_CaCoord":    (H_Water_IF, O_Water_IF_CaCoord), 

        # --- 4. CSH 供体 (H_Hydroxyl/Water_CSH) 跨界/进水 ---
        "H_Hydroxyl_CSH->O_SiO2_Surface":    (H_Hydroxyl_CSH, O_SiO2_Surface),
        "H_Hydroxyl_CSH->O_SiO2_Structural": (H_Hydroxyl_CSH, O_SiO2_Structural),
        "H_Hydroxyl_CSH->O_Water_IF":        (H_Hydroxyl_CSH, O_Water_IF),         
        "H_Hydroxyl_CSH->O_Water_IF_CaCoord":(H_Hydroxyl_CSH, O_Water_IF_CaCoord), 

        "H_Water_CSH->O_SiO2_Surface":       (H_Water_CSH, O_SiO2_Surface),
        "H_Water_CSH->O_SiO2_Structural":    (H_Water_CSH, O_SiO2_Structural),
        "H_Water_CSH->O_Water_IF":           (H_Water_CSH, O_Water_IF),            
        "H_Water_CSH->O_Water_IF_CaCoord":   (H_Water_CSH, O_Water_IF_CaCoord),    

        # --- 5. SiO2 供体 (H_SiO2_Hydroxyl) 跨界/进水 ---
        "H_SiO2_Hydroxyl->O_Hydroxyl_CSH":   (H_SiO2_Hydroxyl, O_Hydroxyl_CSH),
        "H_SiO2_Hydroxyl->O_NBO_CSH":        (H_SiO2_Hydroxyl, O_NBO_CSH),         
        "H_SiO2_Hydroxyl->O_Water_CSH":      (H_SiO2_Hydroxyl, O_Water_CSH),
        "H_SiO2_Hydroxyl->O_Bridging_CSH":   (H_SiO2_Hydroxyl, O_Bridging_CSH),
        "H_SiO2_Hydroxyl->O_Water_IF":       (H_SiO2_Hydroxyl, O_Water_IF),        
        "H_SiO2_Hydroxyl->O_Water_IF_CaCoord":(H_SiO2_Hydroxyl, O_Water_IF_CaCoord)
    })
    # =================================================================
    # [替换结束]
    # =================================================================
    counts = {k: 0 for k in PAIRS.keys()}

    # 准备 donor、acceptor 的窗口内列表
    donors_by_type = {ht: [i for i in ids_by_type.get(ht, []) if i in mask_ids] for ht in donor_H_types}
    acceptors_by_type = {ot: [i for i in ids_by_type.get(ot, []) if i in mask_ids] for ot in acceptor_O_types}

    # 工具：找母氧 Od（先 bonds 再兜底）
    def find_mother_O(hid):
        # 邻接表找 O
        for nb in adj.get(hid, ()):
            if typ.get(nb) in OXYGEN_TYPES:
                return nb
        # 兜底最近 O（≤ NEAR_O_FALLBACK）
        best_r = 1e9; best_id = None
        for ot in OXYGEN_TYPES:
            for oid in ids_by_type.get(ot, ()):
                r = dist_pbc(pos[hid], pos[oid], box)
                if r < best_r:
                    best_r = r; best_id = oid
        if best_id is not None and best_r <= NEAR_O_FALLBACK:
            return best_id
        return None

    # 开始 one-per-donor
    for ht, H_need in [(k, v[0]) for k,v in PAIRS.items()]:
        pass  # 仅为阅读器不报错，无实际意义

    for ht, donor_ids in donors_by_type.items():
        for hid in donor_ids:
            Hp = pos[hid]
            # 找母氧
            Od = find_mother_O(hid)
            if Od is None:
                continue
            Odp = pos[Od]

            # 构建候选受体（窗口内）
            candidates = []  # 元素：(pair_name, Oa_id, angle, r_HA, r_OO)
            for name, (H_need, O_need) in PAIRS.items():
                if ht != H_need:
                    continue
                for Oa in acceptors_by_type.get(O_need, []):
                    if Oa == Od:
                        continue
                    Oap = pos[Oa]
                    # 1) H···O
                    r_HA = dist_pbc(Hp, Oap, box)
                    if r_HA > HBOND_R_OH_MAX:
                        continue
                    # 2) OO（可选）
                    passed_oo = True; r_OO = dist_pbc(Odp, Oap, box)
                    if HBOND_USE_OO_CUT and r_OO > HBOND_OO_MAX:
                        passed_oo = False
                    if not passed_oo:
                        continue
                    # 3) 角度（可选）
                    ang = angle_Od_H_Oa(Odp, Hp, Oap, box) if HBOND_USE_ANGLE else 180.0
                    if HBOND_USE_ANGLE and ang < HBOND_ANGLE_MIN:
                        continue
                    candidates.append((name, Oa, ang, r_HA, r_OO))

            if not candidates:
                continue
            # 选择“最佳受体”
            candidates.sort(key=lambda t: (-t[2], t[3], t[4], t[1]))
            best_pair_name = candidates[0][0]
            counts[best_pair_name] += 1

    # 面积归一化
    A_nm2 = area_xy_nm2(frame_data_aligned)
    counts_per_nm2 = {f"{k}_per_nm2": (v / A_nm2 if A_nm2 > 0 else 0.0) for k, v in counts.items()}
    # 每供体概率归一：pair 按 donor 类型除以窗口内该 donor 类型数量
    pair_donor = {name: H_need for name, (H_need, O_need) in PAIRS.items()}
    donors_total_by_type = {ht: len(donors_by_type.get(ht, [])) for ht in donor_H_types}
    counts_per_donor = {}
    for name, v in counts.items():
        ht = pair_donor[name]
        nD = donors_total_by_type.get(ht, 0)
        counts_per_donor[f"{name}_per_donor"] = (v / nD) if nD > 0 else 0.0
    return counts, counts_per_nm2, counts_per_donor

# =============== 氢键诊断工具（顶层，无缩进） ===============
def diagnose_hbond_coverage(frame_data, donor_types, oxygen_types, fallback_cut=1.25):
    """
    统计 donor H 中能找到母氧 O_d 的覆盖率；并给出 O_d–H 的距离统计。
    """
    pos = frame_data['pos']; typ = frame_data['typ']; ids_by_type = frame_data['ids_by_type']
    adj = frame_data['adj']; box = frame_data['box']

    donors = []
    for ht in donor_types:
        donors.extend(ids_by_type.get(ht, []))

    n_total = len(donors)
    n_with_Od = 0
    dlist = []

    for hid in donors:
        # 先从 Bonds 邻居找 O
        Od = None
        for nb in adj.get(hid, ()):
            if typ.get(nb) in OXYGEN_TYPES:
                Od = nb
                break
        # 兜底：最近 O（≤ fallback_cut）
        if Od is None:
            best_r = 1e9; best_id = None
            for ot in oxygen_types:
                for oid in ids_by_type.get(ot, ()):
                    r = dist_pbc(pos[hid], pos[oid], box)
                    if r < best_r:
                        best_r = r; best_id = oid
            if best_id is not None and best_r <= fallback_cut:
                Od = best_id
                dlist.append(best_r)
        else:
            dlist.append(dist_pbc(pos[hid], pos[Od], box))

        if Od is not None:
            n_with_Od += 1

    return {
        "n_donors": n_total,
        "n_with_Od": n_with_Od,
        "ratio": (n_with_Od / n_total if n_total > 0 else 0.0),
        "Od_H_dist_mean": (sum(dlist)/len(dlist) if dlist else None),
        "Od_H_dist_min": (min(dlist) if dlist else None),
        "Od_H_dist_max": (max(dlist) if dlist else None),
        "sample_size": len(dlist),
    }

def diagnose_hbond_distance_only(frame_data, pair_list, Rmax=2.6):
    """
    仅按 O···H 距离计数（不判角度），看是否有潜在氢键邻近。
    pair_list: [(name, donor_H_type_set, acceptor_O_type_set)]
    """
    pos = frame_data['pos']; typ = frame_data['typ']; ids_by_type = frame_data['ids_by_type']; box = frame_data['box']
    stats = {}
    for name, Hset, Oset in pair_list:
        donors = []
        for ht in Hset:
            donors.extend(ids_by_type.get(ht, []))
        acceptors = []
        for ot in Oset:
            acceptors.extend(ids_by_type.get(ot, []))

        cnt = 0
        for hid in donors:
            Hp = pos[hid]
            for oid in acceptors:
                r = dist_pbc(Hp, pos[oid], box)
                if r <= Rmax:
                    cnt += 1
        stats[name] = {
            "distance_only_count": cnt,
            "n_donors": len(donors),
            "n_acceptors": len(acceptors)
        }
    return stats   
    
    
# =============================================================================
# 分析：RDF 累计（按块平均）
# =============================================================================
class RDFAggregator:
    def __init__(self, pairs, rmax=RDF_R_MAX, dr=RDF_DR):
        """
        pairs: list of (name, A_types_set, B_types_set, is_identical_pair)
               is_identical_pair=True 表示 A 与 B 类型集相同（A==B）
        """
        self.pairs = pairs
        self.rmax = rmax
        self.dr = dr
        self.nbins = int(math.floor(rmax / dr))
        # 每个块的直方图与denom存放
        self.block_hist = []   # list of dict(pair_name -> np.array nbins)
        self.block_denom = []  # list of dict(pair_name -> np.array nbins)（每bin相同，但为统一结构）
        self.current_block_hist = {name: np.zeros(self.nbins, dtype=np.float64) for (name, _, _, _) in pairs}
        self.current_block_denom = {name: np.zeros(self.nbins, dtype=np.float64) for (name, _, _, _) in pairs}
        self.frames_in_current_block = 0


    def update_frame(self, frame_data, V_eff=None):
        box = frame_data['box']; pos = frame_data['pos']; ids_by_type = frame_data['ids_by_type']
        if V_eff is not None:
            V = float(V_eff)
        else:
            V = box_lengths(box)[0] * box_lengths(box)[1] * box_lengths(box)[2]

        for (pair_name, Aset, Bset, identical) in self.pairs:
            A_ids = []
            for t in Aset:
                A_ids.extend(ids_by_type.get(t, []))
            B_ids = []
            for t in Bset:
                B_ids.extend(ids_by_type.get(t, []))

            NA = len(A_ids)
            NB = len(B_ids)
            if NA == 0 or NB == 0:
                continue

            # 期望计数的归一化项（不随 r 变化）
            if not identical:
                denom_scalar = (NA * NB) / V
            else:
                # A==B，遍历i<j
                denom_scalar = (NA * (NA - 1) / 2.0) / V

            # 直方图
            if not identical:
                # A 与 B 可能相交，允许双重循环
                for a in A_ids:
                    pa = pos[a]
                    for b in B_ids:
                        pb = pos[b]
                        # 若 Aset 与 Bset 重叠，会双计 (a,b) 与 (b,a)，这不是问题，因为分母用 NA*NB/V
                        r = dist_pbc(pa, pb, box)
                        bi = bin_index(r, self.rmax, self.dr)
                        if bi >= 0:
                            self.current_block_hist[pair_name][bi] += 1.0
                # 每个 bin 的 denom = 4π r^2 dr × denom_scalar
                r_centers = (np.arange(self.nbins) + 0.5) * self.dr
                shell = 4.0 * math.pi * r_centers * r_centers * self.dr
                self.current_block_denom[pair_name] += shell * denom_scalar
            else:
                # identical: A == B，i<j
                nA = len(A_ids)
                for ii in range(nA):
                    a = A_ids[ii]; pa = pos[a]
                    for jj in range(ii+1, nA):
                        b = A_ids[jj]; pb = pos[b]
                        r = dist_pbc(pa, pb, box)
                        bi = bin_index(r, self.rmax, self.dr)
                        if bi >= 0:
                            self.current_block_hist[pair_name][bi] += 1.0
                r_centers = (np.arange(self.nbins) + 0.5) * self.dr
                shell = 4.0 * math.pi * r_centers * r_centers * self.dr
                self.current_block_denom[pair_name] += shell * denom_scalar

        self.frames_in_current_block += 1

    def close_block(self):
        # 将当前块的 hist/denom 保存，然后重置
        self.block_hist.append({k: v.copy() for k, v in self.current_block_hist.items()})
        self.block_denom.append({k: v.copy() for k, v in self.current_block_denom.items()})
        self.current_block_hist = {name: np.zeros(self.nbins, dtype=np.float64) for (name, _, _, _) in self.pairs}
        self.current_block_denom = {name: np.zeros(self.nbins, dtype=np.float64) for (name, _, _, _) in self.pairs}
        self.frames_in_current_block = 0

    def finalize(self):
        # 若最后一个块未正好填满，也要关闭
        if self.frames_in_current_block > 0:
            self.close_block()
        # 计算块均值与标准误差
        results = {}
        r_centers = (np.arange(self.nbins) + 0.5) * self.dr
        n_blocks = len(self.block_hist)
        for (pair_name, _, _, _) in self.pairs:
            # 逐块 g(r)
            g_blocks = []
            for b in range(n_blocks):
                hist = self.block_hist[b][pair_name]
                denom = self.block_denom[b][pair_name]
                with np.errstate(divide='ignore', invalid='ignore'):
                    g = np.where(denom > 0, hist / denom, 0.0)
                g_blocks.append(g)
            if n_blocks > 0:
                g_blocks_arr = np.stack(g_blocks, axis=0)
                g_mean = np.mean(g_blocks_arr, axis=0)
                # 标准误差（块平均）
                if n_blocks > 1:
                    g_se = np.std(g_blocks_arr, axis=0, ddof=1) / math.sqrt(n_blocks)
                else:
                    g_se = np.zeros_like(g_mean)
            else:
                g_mean = np.zeros(self.nbins); g_se = np.zeros(self.nbins)
            results[pair_name] = {'r': r_centers, 'g_mean': g_mean, 'g_se': g_se, 'n_blocks': n_blocks}
        return results

def guess_r_cut_for_pair(name):
    if "H_" in name and "O_" in name:
        return HBOND_R_OH_MAX
    if "Ca-" in name and "O_" in name:
        return 3.0
    if name.count("O_") >= 2:
        return 3.3
    return 3.5

def count_CN_by_cut_in_window(frame_data_aligned, mask_ids, pairs, rcut_map=None):
    pos = frame_data_aligned['pos']
    box = frame_data_aligned['box']
    ids_by_type = frame_data_aligned['ids_by_type']
    cn = {}
    for (pair_name, Aset, Bset, identical) in pairs:
        A_ids = []
        for t in Aset:
            A_ids.extend([i for i in ids_by_type.get(t, []) if i in mask_ids])
        B_ids = []
        for t in Bset:
            B_ids.extend([i for i in ids_by_type.get(t, []) if i in mask_ids])
        
        if not A_ids or not B_ids:
            cn[pair_name] = 0
            continue
        
        # 确定当前对的截断半径
        rcut = (rcut_map.get(pair_name) if rcut_map and pair_name in rcut_map 
                else guess_r_cut_for_pair(pair_name))
        count = 0
        
        if not identical:
            # 非相同类型对：A-B 双向计数
            for a in A_ids:
                pa = pos[a]
                for b in B_ids:
                    if dist_pbc(pa, pos[b], box) <= rcut:
                        count += 1
        else:
            # 相同类型对：只计 i<j 避免重复
            for ii in range(len(A_ids)):
                pa = pos[A_ids[ii]]
                for jj in range(ii + 1, len(A_ids)):
                    if dist_pbc(pa, pos[A_ids[jj]], box) <= rcut:
                        count += 1
        
        cn[pair_name] = count
    
    # 计算每nm²的配位数（假设area_xy_nm2已定义）
    A_nm2 = area_xy_nm2(frame_data_aligned)
    cn_per_nm2 = {
        f"{k}_per_nm2": (v / A_nm2 if A_nm2 > 0 else 0.0) 
        for k, v in cn.items()
    }
    return cn, cn_per_nm2

# =============================================================================
# 分析：ITZ 密度剖面（z）
# =============================================================================
class DensityZ:
    def __init__(self, types_to_track, dz=0.2):
        self.types = types_to_track
        self.dz = dz
        self.block_bins = []   # list of dict(type -> (z_centers, rho_vals))；这里我们按帧累积，再块平均
        self.current_block_accum = None
        self.current_block_frames = 0

    def _init_accum(self, nbin):
        # 对每个 type：累加 counts（每帧）
        return {t: np.zeros(nbin, dtype=np.float64) for t in self.types}

    def update_frame(self, frame_data):
        box = frame_data['box']; pos = frame_data['pos']; typ = frame_data['typ']
        (xlo, xhi), (ylo, yhi), (zlo, zhi) = box
        Lx, Ly, Lz = xhi - xlo, yhi - ylo, zhi - zlo
        A_xy = Lx * Ly

        nbin = int(math.ceil(Lz / self.dz))
        # bin edges & centers
        z_edges = np.linspace(0.0, Lz, nbin + 1)
        z_centers = 0.5 * (z_edges[:-1] + z_edges[1:])

        if self.current_block_accum is None:
            self.current_block_accum = self._init_accum(nbin)
            self._z_centers = z_centers
            self._A_xy = A_xy
        # 若 bin 数变化（理论不应发生），简单重置
        if len(self._z_centers) != len(z_centers):
            self.current_block_accum = self._init_accum(nbin)
            self._z_centers = z_centers
            self._A_xy = A_xy

        # 按类型统计每帧 count（不除面积，先做计数）
        # 如果需要数密度，可在最终平均时除以 A_xy
        for t in self.types:
            ids = [aid for aid, at in typ.items() if at == t]
            counts = np.zeros(nbin, dtype=np.float64)
            for aid in ids:
                z = pos[aid][2]
                bi = bin_index(z - zlo, Lz, self.dz)  # 这里简化：用 (z - zlo)/dz
                if 0 <= bi < nbin:
                    counts[bi] += 1.0
            self.current_block_accum[t] += counts

        self.current_block_frames += 1

    def close_block(self):
        # 将当前块的平均（每帧平均计数）保存
        if self.current_block_frames == 0:
            # 空块
            self.block_bins.append({t: (self._z_centers.copy(), np.zeros_like(self._z_centers)) for t in self.types})
        else:
            block_entry = {}
            for t in self.types:
                avg_counts = self.current_block_accum[t] / self.current_block_frames
                block_entry[t] = (self._z_centers.copy(), avg_counts.copy())
            self.block_bins.append(block_entry)
        # reset
        self.current_block_accum = None
        self.current_block_frames = 0

    def finalize(self):
        if self.current_block_frames > 0:
            self.close_block()
        # 对每个类型，在各块上取均值±SE
        results = {}
        n_blocks = len(self.block_bins)
        if n_blocks == 0:
            return {t: {'z': np.array([]), 'rho_mean': np.array([]), 'rho_se': np.array([])} for t in self.types}
        for t in self.types:
            zs = self.block_bins[0][t][0]
            mat = np.stack([self.block_bins[b][t][1] for b in range(n_blocks)], axis=0)
            rho_mean = np.mean(mat, axis=0)
            rho_se = np.std(mat, axis=0, ddof=1) / math.sqrt(n_blocks) if n_blocks > 1 else np.zeros_like(rho_mean)
            results[t] = {'z': zs, 'rho_mean': rho_mean, 'rho_se': rho_se, 'n_blocks': n_blocks}
        return results

# =============================================================================
# 用“每块 Axy 平均”把计数剖面转换为质量密度（g/cm^3）
# =============================================================================
def convert_density_counts_blocks_to_mass(dens_block_bins, A_xy_block_means, dz):
    """
    把 DensityZ 的“每块平均计数/帧/ bin”（dens_block_bins）转换为质量密度 g/cm^3：
    - 先对每块：ρ_t^b(z) = N_t^b(z) * MASS_t(amu) * 1.66054 / (Axy_blockb * dz(Å))
    - 再对块：按块平均得到 mean，SE=std(ddof=1)/sqrt(n_blocks)
    返回 dict: {type -> {"z","rho_mean","rho_se","n_blocks"} , TOTAL_MASS 同理}
    """
    if not dens_block_bins:
        return {}
    n_blocks = len(dens_block_bins)
    if len(A_xy_block_means) != n_blocks:
        raise ValueError(f"A_xy_block_means 个数({len(A_xy_block_means)})与密度块数({n_blocks})不一致")
    # 取全类型集合（按块 union）
    types_all = set()
    for b in range(n_blocks):
        types_all |= set(dens_block_bins[b].keys())
    # 取基准 z（用第0块的任一类型）
    any_t0 = next(iter(dens_block_bins[0].keys()))
    z0 = np.asarray(dens_block_bins[0][any_t0][0], dtype=float)

    # 每类型的各块质量密度数组
    mass_blocks_by_type = {t: [] for t in types_all if t in MASS_AMU}
    # 每块的总质量密度数组
    mass_blocks_total = []

    for b in range(n_blocks):
        Axy = float(A_xy_block_means[b])
        total_b = np.zeros_like(z0, dtype=float)

        # 断言各块 z 网格一致（如果不一致，应在 DensityZ 侧保证）
        for t in types_all:
            if t not in MASS_AMU:
                continue
            z_b, counts_b = dens_block_bins[b][t]
            zb = np.asarray(z_b, dtype=float)
            if len(zb) != len(z0):
                raise ValueError("不同块的 z-bin 长度不一致，无法直接块平均。请检查密度分箱设置。")
            counts_b = np.asarray(counts_b, dtype=float)
            factor = MASS_AMU[t] * 1.66053906660 / (Axy * dz)  # amu->g 与 Å^-3->cm^-3 合并为 1.66054
            mass_b = counts_b * factor  # g/cm^3
            mass_blocks_by_type[t].append(mass_b)
            total_b += mass_b

        mass_blocks_total.append(total_b)

    # 汇总为 mean / SE
    results = {}
    for t, blocks in mass_blocks_by_type.items():
        arr = np.stack(blocks, axis=0)  # [n_blocks, n_z]
        mean = np.mean(arr, axis=0)
        se   = (np.std(arr, axis=0, ddof=1) / math.sqrt(n_blocks)) if n_blocks > 1 else np.zeros_like(mean)
        results[t] = {"z": z0, "rho_mean": mean, "rho_se": se, "n_blocks": n_blocks}

    arr_tot = np.stack(mass_blocks_total, axis=0)
    mean_tot = np.mean(arr_tot, axis=0)
    se_tot   = (np.std(arr_tot, axis=0, ddof=1) / math.sqrt(n_blocks)) if n_blocks > 1 else np.zeros_like(mean_tot)
    # 若未在前面给 TOTAL_MASS 定义名字，这里也能工作；只是导出列名用 "type0"
    results.get(0, None)
    results[0] = {"z": z0, "rho_mean": mean_tot, "rho_se": se_tot, "n_blocks": n_blocks}
    return results

def convert_density_counts_to_mass(dens_results, area_xy_ang2, dz):
    """
    把 DensityZ.finalize() 的计数剖面（每bin平均原子数）转换为质量密度 g/cm^3。
    dens_results: {type -> {"z": array, "rho_mean": counts/bin, "rho_se": counts/bin, "n_blocks": int}}
    area_xy_ang2: Å^2（界面窗口的横向面积）
    dz: Å（bin厚度）
    返回：结构同 dens_results，rho_mean/rho_se 单位改为 g/cm^3，并附加 TOTAL_MASS 项
    """
    if not dens_results:
        return {}
    # 取基准 z、块数
    any_t = next(iter(dens_results.keys()))
    z = np.asarray(dens_results[any_t]["z"], dtype=float)
    n_blocks = dens_results[any_t].get("n_blocks", 0)

    mass_results = {}
    total_mean = np.zeros_like(z, dtype=float)
    total_se_sq = np.zeros_like(z, dtype=float)  # 用于 SE 合并：se_total = sqrt(sum(se_i^2))

    for t, res in dens_results.items():
        if t not in MASS_AMU:
            # 未在 MASS_AMU 的类型跳过
            continue
        counts_mean = np.asarray(res["rho_mean"], dtype=float)  # 平均计数/帧/ bin
        counts_se   = np.asarray(res["rho_se"],   dtype=float)  # 块标准误（计数）

        # 因子：ρ(g/cm^3) = N / (A*dz) * mass(amu) * (amu->g) * (Å^-3 -> cm^-3)
        # amu->g = 1.66054e-24，Å^-3 -> cm^-3 乘以 1e24，因此两者合起来就是 × 1.66054
        factor = (MASS_AMU[t] * (AMU_TO_G / ANG3_TO_CM3)) / (area_xy_ang2 * dz)  # = MASS * 1.66054 / (A*dz)

        mass_mean = counts_mean * factor       # g/cm^3
        mass_se   = counts_se   * factor       # g/cm^3（线性缩放）

        mass_results[t] = {"z": z, "rho_mean": mass_mean, "rho_se": mass_se, "n_blocks": n_blocks}

        total_mean += mass_mean
        total_se_sq += (mass_se **2)

    # 组合出总质量密度（逐点相加；SE 用 sqrt(sum(se_i^2)) 近似合并）
    total_se = np.sqrt(total_se_sq)
    mass_results[TOTAL_MASS] = {"z": z, "rho_mean": total_mean, "rho_se": total_se, "n_blocks": n_blocks}
    return mass_results

# =============================================================================
# 块平均工具（用于化学键/氢键计数）
# =============================================================================
def block_average_frames(frame_series, block_size):
    """
    frame_series: list of dict(name -> value)
    返回：mean_dict, se_dict, block_means (list of dict)
    """
    n = len(frame_series)
    blocks = []
    i = 0
    while i < n:
        chunk = frame_series[i:i+block_size]
        if len(chunk) == 0:
            break
        # 该块内对每个键名求均值
        keys = chunk[0].keys()
        block_mean = {}
        for k in keys:
            vals = [d[k] for d in chunk]
            block_mean[k] = sum(vals) / len(vals)
        blocks.append(block_mean)
        i += block_size

    # 全局均值与SE
    mean_dict = {}
    se_dict = {}
    if len(blocks) > 0:
        keys = blocks[0].keys()
        for k in keys:
            arr = np.array([b[k] for b in blocks], dtype=np.float64)
            mean_dict[k] = float(np.mean(arr))
            if len(blocks) > 1:
                se_dict[k] = float(np.std(arr, ddof=1) / math.sqrt(len(blocks)))
            else:
                se_dict[k] = 0.0
    return mean_dict, se_dict, blocks



# =============================================================================
# 主分析流程
# =============================================================================
def main():
    frames = list(range(START_FRAME, START_FRAME + N_FRAMES * FRAME_STEP, FRAME_STEP))
    print(f"准备处理 {len(frames)} 帧，BLOCK_SIZE={BLOCK_SIZE}")

    bonds_per_frame = []
    oh_bonds_per_frame = []
    hbonds_per_frame = []
    cn_per_frame = []
    iface_metrics = [] 
    first_counts_by_type = None

    A_xy_block_values = [] # 当前块内每帧的 Axy（Å^2）
    A_xy_block_means = [] # 每块的 Axy 平均（Å^2）
    # 存每帧的 z_csh_top, z_sio2_bot, z0, deq

    # RDF 对定义（name, A_type_set, B_type_set, identical_flag）
    # =========================================================================
    # 1. RDF 配对定义 (全面覆盖 42 对 + Total 统计)
    # =========================================================================
    
    # 定义辅助集合变量，方便阅读
    SET_O_IF_ALL = {O_Water_IF, O_Water_IF_CaCoord} # 界面水总氧
    SET_O_CSH_ALL = {O_Bridging_CSH, O_Water_CSH, O_Hydroxyl_CSH, O_NBO_CSH} # CSH侧总氧

    rdf_pairs = []

    # --- Group 1: Ca 配位 (含 IF 细分与汇总) ---
    rdf_pairs.extend([
        ("Ca-O_SiO2_Structural",       {Ca}, {O_SiO2_Structural},    False), # 1
        ("Ca-O_SiO2_Surface",          {Ca}, {O_SiO2_Surface},       False), # 2
        ("Ca-O_Water_IF",              {Ca}, {O_Water_IF},           False), # 3
        ("Ca-O_Water_IF_CaCoord",      {Ca}, {O_Water_IF_CaCoord},   False), # 4
        ("Ca-O_Water_IF_total",        {Ca}, SET_O_IF_ALL,           False), # [Total]
    ])

    # --- Group 2: CSH_Si 跨界面相互作用 ---
    rdf_pairs.extend([
        ("CSH_Si-O_SiO2_Structural",   {CSH_Si}, {O_SiO2_Structural},    False), # 5
        ("CSH_Si-O_SiO2_Surface",      {CSH_Si}, {O_SiO2_Surface},       False), # 6
        ("CSH_Si-O_Water_IF",          {CSH_Si}, {O_Water_IF},           False), # 7
        ("CSH_Si-O_Water_IF_CaCoord",  {CSH_Si}, {O_Water_IF_CaCoord},   False), # 8
        ("CSH_Si-O_Water_IF_total",    {CSH_Si}, SET_O_IF_ALL,           False), # [Total]
    ])

    # --- Group 3: Si_Bulk (SiO2体相Si) 与 CSH/IF 氧 ---
    rdf_pairs.extend([
        ("Si_Bulk-O_Bridging_CSH",     {Si_Bulk}, {O_Bridging_CSH},     False), # 9
        ("Si_Bulk-O_Water_CSH",        {Si_Bulk}, {O_Water_CSH},        False), # 10
        ("Si_Bulk-O_Hydroxyl_CSH",     {Si_Bulk}, {O_Hydroxyl_CSH},     False), # 11
        ("Si_Bulk-O_NBO_CSH",          {Si_Bulk}, {O_NBO_CSH},          False), # 12
        ("Si_Bulk-O_Water_IF",         {Si_Bulk}, {O_Water_IF},         False), # 13
        ("Si_Bulk-O_Water_IF_CaCoord", {Si_Bulk}, {O_Water_IF_CaCoord}, False), # 14
        ("Si_Bulk-O_Water_IF_total",   {Si_Bulk}, SET_O_IF_ALL,         False), # [Total]
    ])

    # --- Group 4: Si_Surface (SiO2表面Si) 与 CSH/IF 氧 ---
    rdf_pairs.extend([
        ("Si_Surface-O_Bridging_CSH",    {Si_Surface}, {O_Bridging_CSH},     False), # 15
        ("Si_Surface-O_Water_CSH",       {Si_Surface}, {O_Water_CSH},        False), # 16
        ("Si_Surface-O_Hydroxyl_CSH",    {Si_Surface}, {O_Hydroxyl_CSH},     False), # 17
        ("Si_Surface-O_NBO_CSH",         {Si_Surface}, {O_NBO_CSH},          False), # 18
        ("Si_Surface-O_Water_IF",        {Si_Surface}, {O_Water_IF},         False), # 19
        ("Si_Surface-O_Water_IF_CaCoord",{Si_Surface}, {O_Water_IF_CaCoord}, False), # 20
        ("Si_Surface-O_Water_IF_total",  {Si_Surface}, SET_O_IF_ALL,         False), # [Total]
    ])

    # --- Group 5: H_Water_CSH (CSH层间水氢) 跨界 ---
    rdf_pairs.extend([
        ("H_Water_CSH-O_SiO2_Structural",    {H_Water_CSH}, {O_SiO2_Structural},    False), # 21
        ("H_Water_CSH-O_SiO2_Surface",       {H_Water_CSH}, {O_SiO2_Surface},       False), # 22
        ("H_Water_CSH-O_Water_IF",           {H_Water_CSH}, {O_Water_IF},           False), # 23
        ("H_Water_CSH-O_Water_IF_CaCoord",   {H_Water_CSH}, {O_Water_IF_CaCoord},   False), # 24
        ("H_Water_CSH-O_Water_IF_total",     {H_Water_CSH}, SET_O_IF_ALL,           False), # [Total]
    ])

    # --- Group 6: H_Hydroxyl_CSH (CSH羟基氢) 跨界 ---
    rdf_pairs.extend([
        ("H_Hydroxyl_CSH-O_SiO2_Structural",  {H_Hydroxyl_CSH}, {O_SiO2_Structural},    False), # 25
        ("H_Hydroxyl_CSH-O_SiO2_Surface",     {H_Hydroxyl_CSH}, {O_SiO2_Surface},       False), # 26
        ("H_Hydroxyl_CSH-O_Water_IF",         {H_Hydroxyl_CSH}, {O_Water_IF},           False), # 27
        ("H_Hydroxyl_CSH-O_Water_IF_CaCoord", {H_Hydroxyl_CSH}, {O_Water_IF_CaCoord},   False), # 28
        ("H_Hydroxyl_CSH-O_Water_IF_total",   {H_Hydroxyl_CSH}, SET_O_IF_ALL,           False), # [Total]
    ])

    # --- Group 7: H_SiO2_Hydroxyl (SiO2羟基氢) 跨界 ---
    rdf_pairs.extend([
        ("H_SiO2_Hydroxyl-O_Bridging_CSH",      {H_SiO2_Hydroxyl}, {O_Bridging_CSH},     False), # 29
        ("H_SiO2_Hydroxyl-O_Water_CSH",         {H_SiO2_Hydroxyl}, {O_Water_CSH},        False), # 30
        ("H_SiO2_Hydroxyl-O_Hydroxyl_CSH",      {H_SiO2_Hydroxyl}, {O_Hydroxyl_CSH},     False), # 31
        ("H_SiO2_Hydroxyl-O_NBO_CSH",           {H_SiO2_Hydroxyl}, {O_NBO_CSH},          False), # 32
        ("H_SiO2_Hydroxyl-O_Water_IF",          {H_SiO2_Hydroxyl}, {O_Water_IF},         False), # 33
        ("H_SiO2_Hydroxyl-O_Water_IF_CaCoord",  {H_SiO2_Hydroxyl}, {O_Water_IF_CaCoord}, False), # 34
        ("H_SiO2_Hydroxyl-O_Water_IF_total",    {H_SiO2_Hydroxyl}, SET_O_IF_ALL,         False), # [Total]
        # 补充：SiO2 羟基到 CSH 所有氧的并集（观察总氢键倾向）
        ("H_SiO2_Hydroxyl-O_CSH_all",           {H_SiO2_Hydroxyl}, SET_O_CSH_ALL,        False), 
    ])

    # --- Group 8: H_Water_IF (界面水氢) 向上/向下 ---
    rdf_pairs.extend([
        ("H_Water_IF-O_Bridging_CSH",      {H_Water_IF}, {O_Bridging_CSH},     False), # 35
        ("H_Water_IF-O_Water_CSH",         {H_Water_IF}, {O_Water_CSH},        False), # 36
        ("H_Water_IF-O_Hydroxyl_CSH",      {H_Water_IF}, {O_Hydroxyl_CSH},     False), # 37
        ("H_Water_IF-O_NBO_CSH",           {H_Water_IF}, {O_NBO_CSH},          False), # 38
        ("H_Water_IF-O_SiO2_Structural",   {H_Water_IF}, {O_SiO2_Structural},  False), # 39
        ("H_Water_IF-O_SiO2_Surface",      {H_Water_IF}, {O_SiO2_Surface},     False), # 40
        ("H_Water_IF-O_Water_IF",          {H_Water_IF}, {O_Water_IF},         False), # 41 (Free-Free)
        ("H_Water_IF-O_Water_IF_CaCoord",  {H_Water_IF}, {O_Water_IF_CaCoord}, False), # 42 (Free-Bound)
        # 补充：界面水内部互作总览 (H_IF -> O_IF_Total)
        ("H_Water_IF-O_Water_IF_total",    {H_Water_IF}, SET_O_IF_ALL,         False),
    ])

    # --- Group 9: 诊断性对 (未成键氢检查) ---
    # 这些对通常不用于主文图表，但用于检查结构合理性
    rdf_pairs.extend([
        ("H_Unbonded_CSH-O_SiO2_Surface",    {H_Unbonded_CSH}, {O_SiO2_Surface},    False),
        ("H_Unbonded_CSH-O_SiO2_Structural", {H_Unbonded_CSH}, {O_SiO2_Structural}, False),
        ("H_Unbonded_SiO2-O_Hydroxyl_CSH",   {H_Unbonded_SiO2}, {O_Hydroxyl_CSH},   False),
        ("H_Unbonded_CSH-O_Hydroxyl_CSH",    {H_Unbonded_CSH}, {O_Hydroxyl_CSH},    False),
    ])
    
    # 去重（防止上面复制粘贴时有重复项）
    rdf_pairs_unique = []
    seen_names = set()
    for p in rdf_pairs:
        if p[0] not in seen_names:
            rdf_pairs_unique.append(p)
            seen_names.add(p[0])
    rdf_pairs = rdf_pairs_unique

    # =========================================================================
    # 2. 预览作图分组标签 (RDF_PLOT_GROUPS)
    # =========================================================================
    # F3: Ca 相关
    # F6: Si 跨界面 (CSH_Si/Si_Bulk/Si_Surf -> O_Opposite)
    # F7: H 跨界面 (H_CSH/H_SiO2 -> O_Opposite)
    # F8: IF 水相关 (H_IF -> Both Sides)
    
    RDF_PLOT_GROUPS = {
        # --- F3: Ca 配位 ---
        "Ca-O_SiO2_Surface":        "F3-1",
        "Ca-O_SiO2_Structural":     "F3-2",
        "Ca-O_Water_IF":            "F3-3",
        "Ca-O_Water_IF_CaCoord":    "F3-4",
        "Ca-O_Water_IF_total":      "F3-All",

        # --- F6: Si 骨架跨界面相互作用 ---
        # CSH_Si -> SiO2/IF
        "CSH_Si-O_SiO2_Surface":       "F6-CSH-1",
        "CSH_Si-O_Water_IF_total":     "F6-CSH-2",
        # Si_Surface -> CSH/IF
        "Si_Surface-O_Hydroxyl_CSH":   "F6-SiSurf-1",
        "Si_Surface-O_Water_CSH":      "F6-SiSurf-2",
        "Si_Surface-O_NBO_CSH":        "F6-SiSurf-3",
        "Si_Surface-O_Water_IF_total": "F6-SiSurf-4",
        # Si_Bulk (较少见，但也列出)
        "Si_Bulk-O_Water_CSH":         "F6-SiBulk-1",
        "Si_Bulk-O_Water_IF_total":    "F6-SiBulk-2",

        # --- F7: 羟基/层间水氢 跨界面氢键 ---
        # H_SiO2 -> CSH
        "H_SiO2_Hydroxyl-O_Hydroxyl_CSH": "F7-SiH-1",
        "H_SiO2_Hydroxyl-O_Water_CSH":    "F7-SiH-2",
        "H_SiO2_Hydroxyl-O_NBO_CSH":      "F7-SiH-3",
        "H_SiO2_Hydroxyl-O_Water_IF_total":"F7-SiH-4",
        "H_SiO2_Hydroxyl-O_CSH_all":      "F7-SiH-All",
        # H_CSH (Hydroxyl) -> SiO2
        "H_Hydroxyl_CSH-O_SiO2_Surface":    "F7-CH-1",
        "H_Hydroxyl_CSH-O_Water_IF_total":  "F7-CH-2",
        # H_CSH (Water) -> SiO2
        "H_Water_CSH-O_SiO2_Surface":       "F7-CW-1",
        "H_Water_CSH-O_Water_IF_total":     "F7-CW-2",

        # --- F8: 界面水 (IF Water) 氢键网络 ---
        # H_IF -> CSH
        "H_Water_IF-O_Hydroxyl_CSH":    "F8-IF-CSH-1",
        "H_Water_IF-O_NBO_CSH":         "F8-IF-CSH-2",
        "H_Water_IF-O_Water_CSH":       "F8-IF-CSH-3",
        # H_IF -> SiO2
        "H_Water_IF-O_SiO2_Surface":    "F8-IF-Si-1",
        "H_Water_IF-O_SiO2_Structural": "F8-IF-Si-2",
        # H_IF -> IF (Self)
        "H_Water_IF-O_Water_IF":        "F8-IF-IF-1",
        "H_Water_IF-O_Water_IF_CaCoord":"F8-IF-IF-2",
        "H_Water_IF-O_Water_IF_total":  "F8-IF-Total",
        
        # --- F9: 诊断组 ---
        "H_Unbonded_CSH-O_SiO2_Surface": "F9-Diag-1",
    }

    rdf_agg = RDFAggregator(rdf_pairs, rmax=RDF_R_MAX, dr=RDF_DR)

    # ITZ 密度剖面
    density_types = [t for t in TYPE_NAME.keys() if isinstance(t, int) and t != 0]
    densZ = DensityZ(density_types, dz=0.2)

    # 逐帧处理
    frames_in_block = 0
    A_xy_ang2_last = None
    # 按配对分组存储NB块均值列表：键为配对（pair），值为该配对对应的多个NB块均值（[NB_block_mean, ...]）
    nb_block_means_by_pair = defaultdict(list)

    # 存储所有Veff块均值的列表：每个元素为一个Veff块均值（[Veff_block_mean, ...]）
    veff_block_means = []

    # 按配对累计当前块的NB值：键为配对（pair），值为该配对在当前块中逐帧累加的NB累计值（float类型）
    _nb_accum_by_pair = defaultdict(float)

    # 存储当前块中每帧的Veff值列表：每个元素为单帧的Veff值（单位：Å³）
    _veff_block_values = []

    # 与frames_in_block同步使用，记录当前RDF（径向分布函数）相关块的已累计帧数
    _frames_in_block_for_rdf = 0
    for idx, fr in enumerate(frames):
        path = DATA_PATTERN.format(i=fr)
        print(f"\n--- 读取帧 {fr}: {os.path.basename(path)} ---")
        try:
            # 1) 读取与标记
            data = parse_lammps_data(path)
            # 2) 界面定位 + half_w（只依赖 deq 和常数）
            z_csh_top, z_sio2_bot, z0, deq = compute_interface_markers(data, IF_Q_CSH_TOP, IF_Q_SIO2_BOT)
            zlo_abs = data['box'][2][0]
            half_w = (deq * 0.5) + IF_WINDOW_HALF

            # 3) 对齐坐标、面积缓存
            data_aligned = make_aligned_frame(data, z0)
            (xlo, xhi), (ylo, yhi), _ = data_aligned['box']
            A_xy_ang2_last = (xhi - xlo) * (yhi - ylo)  # Å^2
            A_xy_block_values.append(A_xy_ang2_last)

            # 4) 窗口掩码（用刚算好的 half_w）
            mask_ids = window_mask_ids(data_aligned, half_w)

            # 5) 覆盖率自检（必须在 append iface_metrics 之前）
            all_csh = data_aligned['ids_by_type'].get(CSH_Si, [])
            n_csh_total = len(all_csh)
            n_csh_in_win = sum(1 for i in all_csh if i in mask_ids)
            frac_csh_in_window = (n_csh_in_win / n_csh_total) if n_csh_total > 0 else 0.0
            n_Ca_in_window = len([i for i in data_aligned['ids_by_type'].get(Ca, []) if i in mask_ids])

            def median_anchor(z_arr, side, frac=0.30):
                """
                计算相内锚点的z坐标（对齐坐标）
                取相内最内侧frac比例原子的z坐标中位数作为锚点
                参数:
                    z_arr: 原子的z坐标数组（对齐后）
                    side: 相类型 ("CSH" 或其他，对应SiO2)
                    frac: 取最内侧原子的比例（默认0.30）
                返回:
                    锚点的z坐标（对齐后），数组为空时返回None
                """
                if z_arr.size == 0:
                    return None
                z_sorted = np.sort(z_arr)
                k = max(1, int(round(frac * len(z_sorted))))
                if side == "CSH":  # 更深入CSH体内 → 对齐坐标更负
                    subset = z_sorted[:k]
                else:  # 更深入SiO2体内 → 对齐坐标更正（side="Si"）
                    subset = z_sorted[-k:]
                return float(np.median(subset))


            # —— 锚点计算：各相最内侧30% Si 的中位数（对齐坐标 z' = z - z0）——
            # 1. 获取对应类型原子的ID
            csh_ids = data_aligned['ids_by_type'].get(CSH_Si, [])
            si_ids = []
            for t in (Si_Bulk, Si_Surface):
                si_ids.extend(data_aligned['ids_by_type'].get(t, []))

            # 2. 提取对齐后的z坐标（z' = z - z0）
            z_csh_al = np.array([data_aligned['pos'][i][2] for i in csh_ids], dtype=float)
            z_si_al = np.array([data_aligned['pos'][i][2] for i in si_ids], dtype=float)

            # 3. 计算锚点z坐标（对齐后）
            anchor_csh_z = median_anchor(z_csh_al, side="CSH", frac=0.30)  # 负值（更深入CSH）
            anchor_sio2_z = median_anchor(z_si_al, side="Si", frac=0.30)   # 正值（更深入SiO2）

            # 4. 计算对齐后的表面z坐标（理论上为 ±deq/2）
            z_csh_top_al = (z_csh_top - z0)    # CSH表面对齐坐标（-deq/2）
            z_sio2_bot_al = (z_sio2_bot - z0)  # SiO2表面对齐坐标（+deq/2）

            # 5. 计算表面到锚点的体内深度（>0 表示从表面向体内的距离）
            d_csh_surface_to_anchor = None if anchor_csh_z is None else (z_csh_top_al - anchor_csh_z)
            d_sio2_surface_to_anchor = None if anchor_sio2_z is None else (anchor_sio2_z - z_sio2_bot_al)

            # 6) 记录 iface_metrics（此时 half_w_used 与 frac_* 已有）
            iface_metrics.append({
                "frame": int(fr),
                "z_csh_top": z_csh_top, "z_sio2_bot": z_sio2_bot, "z0": z0, "deq": deq,
                "z0_rel": z0 - zlo_abs, "z_csh_top_rel": z_csh_top - zlo_abs, "z_sio2_bot_rel": z_sio2_bot - zlo_abs,
                "IF_WINDOW_HALF": IF_WINDOW_HALF,
                "half_w_used": half_w,
                "frac_CSH_in_window": frac_csh_in_window,
                "n_Ca_in_window": n_Ca_in_window,
                # 新增：锚点与表面→锚点距离（对齐坐标）
                "anchor_CSH_z_aligned": "" if anchor_csh_z is None else anchor_csh_z,
                "anchor_SiO2_z_aligned": "" if anchor_sio2_z is None else anchor_sio2_z,
                "d_CSH_surface_to_anchor": "" if d_csh_surface_to_anchor is None else d_csh_surface_to_anchor,
                "d_SiO2_surface_to_anchor": "" if d_sio2_surface_to_anchor is None else d_sio2_surface_to_anchor,
            })

            # 7) 体积归一参数（A_nm2 / V_nm3）
            A_nm2 = area_xy_nm2(data_aligned)                  # nm^2
            thickness_nm = 2.0 * half_w * 0.1                  # Å -> nm
            V_nm3 = A_nm2 * thickness_nm if thickness_nm > 0 else 0.0

            # 8) 窗口过滤、计数 per_nm2/per_nm3、RDF 用 V_eff
            data_win_for_bonds = filter_frame_for_window(data_aligned, mask_ids, filter_bonds=APPLY_WINDOW_TO_BONDS)
            data_win_for_rdf   = filter_frame_for_window(data_aligned, mask_ids, filter_bonds=False)

            # 化学键
            c_bonds = count_chemical_bonds(data_win_for_bonds)

            # 每 Ca 归一：该类 Ca–O 计数 / 窗口内 Ca 数
            c_bonds_per_Ca = {
                f"{k}_per_Ca": (c_bonds.get(k, 0) / n_Ca_in_window) if n_Ca_in_window > 0 else 0.0
                for k in CA_O_KEYS
            }

            CN_Ca_total_per_Ca = (sum(c_bonds.get(k, 0) for k in CA_O_KEYS) / n_Ca_in_window) if n_Ca_in_window > 0 else 0.0

            bonds_per_frame.append({
                **c_bonds,
                **{f"{k}_per_nm2": (v / A_nm2 if A_nm2 > 0 else 0.0) for k, v in c_bonds.items()},
                **{f"{k}_per_nm3": (v / V_nm3 if V_nm3 > 0 else 0.0) for k, v in c_bonds.items()},
                **c_bonds_per_Ca,
                "CN_Ca_total_per_Ca": CN_Ca_total_per_Ca, # 可选
            })


            # 共价 OH
            oh_counts = count_oh_bonds(data_win_for_bonds)
            oh_bonds_per_frame.append({
                **oh_counts,
                **{f"{k}_per_nm2": (v / A_nm2 if A_nm2 > 0 else 0.0) for k, v in oh_counts.items()},
                **{f"{k}_per_nm3": (v / V_nm3 if V_nm3 > 0 else 0.0) for k, v in oh_counts.items()},
            })

            # 氢键 one-per-donor（窗口内）
            hb_counts_raw, hb_counts_per_nm2, hb_counts_per_donor = count_hbonds_one_per_donor_window(data_aligned, mask_ids)
            hb_counts_per_nm3 = {f"{k}_per_nm3": (v / V_nm3 if V_nm3 > 0 else 0.0) for k, v in hb_counts_raw.items()}
            hbonds_per_frame.append({**hb_counts_raw, **hb_counts_per_nm2, **hb_counts_per_nm3, **hb_counts_per_donor})

            # RDF（窗口内；用窗口体积做归一）
            V_eff_ang3 = A_xy_ang2_last * (2.0 * half_w)       # Å^3
            # 记录当前帧的窗口体积（单位：Å³），用于后续 RDF 计算中的体密度 ρ_B
            _veff_block_values.append(V_eff_ang3)

            # 为每对 RDF 统计本帧的受体原子数 NB（窗口内 B 集合所有原子类型的数量之和）
            for (pair_name, Aset, Bset, identical) in rdf_pairs:
                NB = 0  # 初始化当前配对本帧的 NB 值（窗口内 B 集合原子总数）
                # 遍历当前配对的所有 B 原子类型（t 为原子类型标识）
                for t in Bset:
                    # 获取该原子类型对应的所有原子 ID（无则返回空列表）
                    ids = data_aligned['ids_by_type'].get(t, [])
                    # 统计该类型中处于有效窗口内的原子数（ID 在 mask_ids 中即为有效）
                    NB += sum(1 for i in ids if i in mask_ids)
                # 累加当前配对在本块中的 NB 总数（逐帧累计）
                _nb_accum_by_pair[pair_name] += NB

            # 累计当前 RDF 块的已处理帧数（与 frames_in_block 同步）
            _frames_in_block_for_rdf += 1
            rdf_agg.update_frame(data_win_for_rdf, V_eff=V_eff_ang3)

            # CN（窗口内，几何阈值直计）
            cn_raw, cn_norm = count_CN_by_cut_in_window(data_aligned, mask_ids, rdf_pairs, rcut_map=None)
            cn_per_frame.append({**cn_raw, **cn_norm})

            # 密度（对齐坐标）
            densZ.update_frame(data_aligned)

            # 收尾关块
            frames_in_block += 1
            if frames_in_block == BLOCK_SIZE:
                rdf_agg.close_block()
                # —— 计算 RDF 所需的 NB 块均值与 Veff 块均值 ——
                if _frames_in_block_for_rdf > 0:
                    # 计算当前块的 Veff 均值（窗口体积均值），转换为 float 类型后存入列表
                    veff_block_means.append(float(np.mean(_veff_block_values)))
                    
                    # 遍历所有 RDF 配对，计算每个配对的 NB 块均值（累计 NB 数 / 块内帧数）
                    for name in _nb_accum_by_pair.keys():
                        nb_block_mean = _nb_accum_by_pair[name] / _frames_in_block_for_rdf
                        nb_block_means_by_pair[name].append(nb_block_mean)

                # 重置当前块的累计变量，为下一个块的统计做准备
                _nb_accum_by_pair.clear()       # 清空配对的 NB 累计值
                _veff_block_values.clear()      # 清空当前块的 Veff 逐帧值列表
                _frames_in_block_for_rdf = 0    # 重置当前块的已处理帧数计数器
                densZ.close_block()
                A_xy_block_means.append(float(np.mean(A_xy_block_values)))
                A_xy_block_values.clear()
                frames_in_block = 0
        except Exception as e:
            print(f"  跳过：读取失败 - {e}")
            continue

        # 仅在第0帧做诊断输出
        if idx == 0:
            first_counts_by_type = {t: len(data['ids_by_type'].get(t, [])) for t in density_types}
            # 1) 母氧覆盖率诊断
            donor_H_types = {H_SiO2_Hydroxyl, H_Hydroxyl_CSH, H_Water_IF, H_Water_CSH}
            cov = diagnose_hbond_coverage(
                frame_data=data,
                donor_types=donor_H_types,
                oxygen_types=OXYGEN_TYPES,
                fallback_cut=NEAR_O_FALLBACK
            )
            print(
                f"HBOND COVERAGE @frame {fr}: donors={cov['n_donors']}, "
                f"with_Od={cov['n_with_Od']} ({cov['ratio']*100:.1f}%), "
                f"Od–H mean={cov['Od_H_dist_mean']}, "
                f"min={cov['Od_H_dist_min']}, max={cov['Od_H_dist_max']} "
                f"(sample={cov['sample_size']})"
            )

            # 2) 仅按距离（不判角度）的快速计数
            distance_pairs = [
                ("H_Water_IF->O_SiO2_Surface", {H_Water_IF}, {O_SiO2_Surface}),
                ("H_Water_IF->O_Hydroxyl_CSH", {H_Water_IF}, {O_Hydroxyl_CSH}),
                ("H_SiO2_Hydroxyl->O_Hydroxyl_CSH", {H_SiO2_Hydroxyl}, {O_Hydroxyl_CSH}),
                ("H_SiO2_Hydroxyl->O_Water_IF", {H_SiO2_Hydroxyl}, {O_Water_IF}),
                ("H_SiO2_Hydroxyl->O_Water_IF_CaCoord", {H_SiO2_Hydroxyl}, {O_Water_IF_CaCoord}),
            ]
            dist_only = diagnose_hbond_distance_only(
                frame_data=data,
                pair_list=distance_pairs,
                Rmax=HBOND_R_OH_MAX
            )
            for name, v in dist_only.items():
                print(
                    f"DIST-ONLY @frame {fr}: {name}: within {HBOND_R_OH_MAX} Å = "
                    f"{v['distance_only_count']} (donors={v['n_donors']}, acceptors={v['n_acceptors']})"
                )

            # 3) 原有 1:1 自检
            n_Osurf = len(data['ids_by_type'].get(O_SiO2_Surface, []))
            n_Hsurf = len(data['ids_by_type'].get(H_SiO2_Hydroxyl, []))
            print(f"  自检: O_SiO2_Surface={n_Osurf}, H_SiO2_Hydroxyl={n_Hsurf} (应接近1:1)")


        # 自检：N(O_SiO2_Surface) vs N(H_SiO2_Hydroxyl)
        n_Oh = len(data['ids_by_type'].get(O_SiO2_Surface, []))
        n_Ho = len(data['ids_by_type'].get(H_SiO2_Hydroxyl, []))
        if idx == 0:
            print(f"  自检: O_SiO2_Surface={n_Oh}, H_SiO2_Hydroxyl={n_Ho} (应接近1:1)")

    # 若最后一块未满，仍需关闭
    if frames_in_block > 0:
        rdf_agg.close_block()
        densZ.close_block()
    # 计算 RDF 所需的 NB 块均值与 Veff 块均值，并重置累计变量
    if _frames_in_block_for_rdf > 0:
        # 计算当前块的 Veff 均值（窗口体积均值），转换为 float 后存入列表
        veff_block_means.append(float(np.mean(_veff_block_values)))
        
        # 遍历所有 RDF 配对，计算每个配对的 NB 块均值（累计 NB 数 / 块内帧数）并存储
        for name in _nb_accum_by_pair.keys():
            nb_block_means_by_pair[name].append(_nb_accum_by_pair[name] / _frames_in_block_for_rdf)
        
        # 重置当前块的累计变量，为下一个块统计做准备
        _nb_accum_by_pair.clear(); _veff_block_values.clear(); _frames_in_block_for_rdf = 0
    if A_xy_block_values:
        A_xy_block_means.append(float(np.mean(A_xy_block_values)))
        A_xy_block_values.clear()
    # ---------------- 块平均 ----------------
    # 化学键
    bond_mean, bond_se, bond_blocks = block_average_frames(bonds_per_frame, BLOCK_SIZE)
    # 氢键
    hbond_mean, hbond_se, hbond_blocks = block_average_frames(hbonds_per_frame, BLOCK_SIZE)
    # RDF
    rdf_results = rdf_agg.finalize()
    out_dir = os.path.join(WORK_PATH, "analysis_out")
    os.makedirs(out_dir, exist_ok=True)


    # from 对应模块 import cumtrapz_with_zero, smooth_sg, find_first_two_minima
    # 已定义常量：CN_SG_WINDOW, CN_SG_POLY, VALLEY1_RMAX, VALLEY2_RMAX, RDF_PLOT_GROUPS



    # 调用函数：使用RDF_PLOT_GROUPS作为配对标签映射
    compute_rdf_cn_for_all_pairs(
        rdf_results, nb_block_means_by_pair, veff_block_means,
        out_dir, pair_tag_map=RDF_PLOT_GROUPS
    )
    # ITZ
    dens_results = densZ.finalize()
    #将“计数版”密度按块 Axy 转为质量密度 g/cm^3
    dens_mass_results = convert_density_counts_blocks_to_mass(densZ.block_bins, A_xy_block_means, densZ.dz)
    #if A_xy_ang2_last is None:
    #    raise RuntimeError("A_xy_ang2_last 未记录，无法转换到 g/cm^3。请确认在循环内已赋值。")
    #dens_mass_results = convert_density_counts_to_mass(dens_results, A_xy_ang2_last, densZ.dz)
    #质量密度 CSV 导出（新增文件）
    out_dir = os.path.join(WORK_PATH, "analysis_out")
    dens_types_sorted_mass = sorted(dens_mass_results.keys(), key=lambda x: (x != 0, TYPE_NAME.get(x, f"type{x}")))
    base_z = np.asarray(dens_mass_results[dens_types_sorted_mass[0]]["z"], dtype=float)

    def num_plain_str(x, decimals=10):
        """
        转成不带科学计数法、无前导引号的十进制字符串。
        - 去掉BOM/空白/各种引号
        - 转 float 后用固定小数位 'f' 格式
        - 去掉多余的尾随0和末尾小数点
        """
        s = str(x).strip().lstrip("\ufeff").lstrip("'").lstrip("’").lstrip("‘")
        # 也去掉零宽空白
        s = s.replace("\u200b", "")
        try:
            v = float(s)
        except Exception:
            # 转不了就返回清洗后的原字符串
            return s
        out = f"{v:.{decimals}f}" # 固定小数位，杜绝 e 计数法
        if "." in out:
            out = out.rstrip("0").rstrip(".") # 去尾零
        if out == "" or out == "-0":
            out = "0"
        return out
    
    DEC_Z = 6 # z坐标保留的小数位
    DEC_RHO = 10 # 质量密度保留的小数位

    with open(os.path.join(out_dir, "density_z_mass_combined.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        header = ["z(Angstrom)"]
        for t in dens_types_sorted_mass:
            name = TYPE_NAME.get(t, f"type{t}")
            header += [f"{name}_gcm3_mean", f"{name}_gcm3_se"]
        w.writerow(header)

        for i in range(len(base_z)):
            row = [num_plain_str(base_z[i], DEC_Z)]
            for t in dens_types_sorted_mass:
                row += [
                    num_plain_str(dens_mass_results[t]["rho_mean"][i], DEC_RHO),
                    num_plain_str(dens_mass_results[t]["rho_se"][i],   DEC_RHO),
                ]
            w.writerow(row)
    

    #额外导出一份“Excel文本版”CSV：用不可见字符令 Excel 以文本显示小数
    excelfriendly_path = os.path.join(out_dir, "density_z_mass_combined_excelfriendly.csv")

    def as_excel_text(s):
        # 在前面加 U+200A Hair Space（不可见），让 Excel 识别为文本，但显示内容是纯小数
        return "\u200A" + s

    with open(excelfriendly_path, "w", newline='', encoding="utf-8-sig") as f2:
        w2 = csv.writer(f2, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        header = ["z(Angstrom)"]
        for t in dens_types_sorted_mass:
            name = TYPE_NAME.get(t, f"type{t}")
            header += [f"{name}_gcm3_mean", f"{name}_gcm3_se"]
        w2.writerow(header)
        for i in range(len(base_z)):
            row2 = [as_excel_text(num_plain_str(base_z[i], DEC_Z))]
            for t in dens_types_sorted_mass:
                row2 += [
                    as_excel_text(num_plain_str(dens_mass_results[t]["rho_mean"][i], DEC_RHO)),
                    as_excel_text(num_plain_str(dens_mass_results[t]["rho_se"][i], DEC_RHO)),
                ]
            w2.writerow(row2)

    # Excel 新增质量密度工作表，增添工作表（质量密度）
    try:
        import pandas as pd
        excel_path = os.path.join(out_dir, "analysis_all.xlsx")
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_dens_mass = pd.DataFrame({"z(Angstrom)": base_z})
            for t in dens_types_sorted_mass:
                name = TYPE_NAME.get(t, f"type{t}")
                df_dens_mass[f"{name}_gcm3_mean"] = np.asarray(dens_mass_results[t]["rho_mean"], dtype=float)
                df_dens_mass[f"{name}_gcm3_se"] = np.asarray(dens_mass_results[t]["rho_se"], dtype=float)
            df_dens_mass.to_excel(writer, sheet_name="density_z_mass_combined", index=False)
            # 设定 Excel 显示格式（禁止科学计数法）
            ws = writer.sheets["density_z_mass_combined"]
            n_rows, n_cols = df_dens_mass.shape
            # 第1列 z 用 6 位小数，其余列用 10 位小数
            for col_idx in range(1, n_cols + 1):
                fmt = "0.000000" if col_idx == 1 else "0." + "0"*10
                col_letter = get_column_letter(col_idx)
                for row_idx in range(2, n_rows + 2):  # 跳过表头，数据从第2行开始
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt
        print("已导出: density_z_mass_combined.csv，并写入 Excel: density_z_mass_combined")
    except ImportError:
        print("pandas 未安装，跳过质量密度的 Excel 写入")
        
    
    # 密度核对：验证z轴方向积分值与每帧该类型的平均总数是否一致
    print("[Density sanity check] sum over z-bins ≈ avg count per frame:")

    # 遍历需要核对的类型列表，检查各类型的密度积分结果
    for t in [Ca, CSH_Si, O_SiO2_Surface, O_Hydroxyl_CSH]:
        # 仅处理存在于密度结果中的类型
        if t in dens_results:
            # 计算z轴方向的密度积分总和
            tot = float(np.sum(dens_results[t]["rho_mean"]))
            
            # 若存在初始计数数据，则则同时显示初始帧计数
            if first_counts_by_type is not None:
                print(f" {TYPE_NAME[t]:20s}: {fnum(tot)} (frame0 count={first_counts_by_type.get(t, 0)})")
            else:
                print(f" {TYPE_NAME[t]:20s}: {fnum(tot)}")
        
    
    cn_mean, cn_se, cn_blocks = block_average_frames(cn_per_frame, BLOCK_SIZE)

    oh_mean, oh_se, oh_blocks = block_average_frames(oh_bonds_per_frame, BLOCK_SIZE)

    # ---------------- 导出 CSV ----------------


    per_ca_keys = [k for k in bond_mean.keys() if k.endswith("_per_Ca") or k == "CN_Ca_total_per_Ca"]
    with open(os.path.join(out_dir, "ca_cn_per_ca_block_mean_se.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(["pair", "mean", "se", "n_blocks"])
        n_blocks_bonds = len(bond_blocks)
        for k in per_ca_keys:
            w.writerow([k, fnum(bond_mean.get(k, 0.0)), fnum(bond_se.get(k, 0.0)), int(n_blocks_bonds)])


    # 1) 每帧化学键计数
    bond_keys = list(bonds_per_frame[0].keys()) if bonds_per_frame else []
    with open(os.path.join(out_dir, "bonds_counts_per_frame.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["frame"] + bond_keys)
        for i, fr in enumerate(frames[:len(bonds_per_frame)]):
            row = [int(fr)] + [fnum(bonds_per_frame[i][k]) for k in bond_keys]
            w.writerow(row)

    # 1.1) 化学键块平均
    with open(os.path.join(out_dir, "bonds_counts_block_mean_se.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["pair", "mean", "se", "n_blocks"])
        n_blocks_bonds = len(bond_blocks)
        for k in bond_keys:
            w.writerow([k, fnum(bond_mean.get(k, 0.0)), fnum(bond_se.get(k, 0.0)), int(n_blocks_bonds)])

    # 2) 每帧氢键计数（one-per-donor + per_nm2 已经合并在 hbonds_per_frame 的字典里）
    hbond_keys = list(hbonds_per_frame[0].keys()) if hbonds_per_frame else []
    with open(os.path.join(out_dir, "hbonds_counts_per_frame.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["frame"] + hbond_keys)
        for i, fr in enumerate(frames[:len(hbonds_per_frame)]):
            row = [int(fr)] + [fnum(hbonds_per_frame[i][k]) for k in hbond_keys]
            w.writerow(row)

    # 2.1) 氢键块平均
    with open(os.path.join(out_dir, "hbonds_counts_block_mean_se.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["pair", "mean", "se", "n_blocks"])
        n_blocks_hb = len(hbond_blocks)
        for k in hbond_keys:
            w.writerow([k, fnum(hbond_mean.get(k, 0.0)), fnum(hbond_se.get(k, 0.0)), int(n_blocks_hb)])

    # 3) 每帧 O–H 共价键计数
    oh_keys = list(oh_bonds_per_frame[0].keys()) if oh_bonds_per_frame else []
    with open(os.path.join(out_dir, "oh_counts_per_frame.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["frame"] + oh_keys)
        for i, fr in enumerate(frames[:len(oh_bonds_per_frame)]):
            row = [int(fr)] + [fnum(oh_bonds_per_frame[i][k]) for k in oh_keys]
            w.writerow(row)

    # 3.1) O–H 共价键块平均
    with open(os.path.join(out_dir, "oh_counts_block_mean_se.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(["pair", "mean", "se", "n_blocks"])
        n_blocks_oh = len(oh_blocks)
        for k in oh_keys:
            w.writerow([k, fnum(oh_mean.get(k, 0.0)), fnum(oh_se.get(k, 0.0)), int(n_blocks_oh)])

    # 4) RDF 导出（每对一份 CSV），文件名加标签前缀
    plot_groups_rows = [] # 收集映射表
    for pair_name, res in rdf_results.items():
        r = res['r']; g_mean = res['g_mean']; g_se = res['g_se']; n_blocks_rdf = res['n_blocks']
        tag = RDF_PLOT_GROUPS.get(pair_name, "")
        fname = f"rdf_{(tag + '_') if tag else ''}{pair_name}.csv"
        with open(os.path.join(out_dir, fname), "w", newline='', encoding="utf-8-sig") as f:
            w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
            w.writerow(["r(Angstrom)", "g_mean", "g_se", "n_blocks", "plot_tag"])
            for i in range(len(r)):
                w.writerow([fnum(r[i]), fnum(g_mean[i]), fnum(g_se[i]), int(n_blocks_rdf), tag])
        plot_groups_rows.append([pair_name, tag])

    #写一份“RDF 作图分组”映射表，便于查阅/自动作图
    with open(os.path.join(out_dir, "rdf_plot_groups.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(["pair_name", "plot_tag"])
        for row in plot_groups_rows:
            w.writerow(row)
    
    
    
        

    if dens_results:
        # 统一 z 轴
        any_t = next(iter(dens_results.keys()))
        base_z = np.asarray(dens_results[any_t]["z"], dtype=float)
        
        # 类型按名称排序，列更整齐
        dens_types_sorted = sorted(dens_results.keys(), key=lambda x: TYPE_NAME.get(x, f"type{x}"))
        
        dens_csv_path = os.path.join(out_dir, "density_z_combined.csv")
        with open(dens_csv_path, "w", newline='', encoding="utf-8-sig") as f:
            w = csv.writer(f, lineterminator="\r\n")
            
            # 表头：z + 每个类型的 mean 与 se
            header = ["z(Angstrom)"]
            for t in dens_types_sorted:
                name = TYPE_NAME.get(t, f"type{t}")
                header += [f"{name}_mean", f"{name}_se"]
            w.writerow(header)
            
            # 逐行写数据
            for i in range(len(base_z)):
                row = [fnum(base_z[i])]
                for t in dens_types_sorted:
                    row += [fnum(dens_results[t]["rho_mean"][i]), fnum(dens_results[t]["rho_se"][i])]
                w.writerow(row)

    #5) 界面指标 CSV（每帧）
    with open(os.path.join(out_dir, "interface_metrics_per_frame.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow([
        "frame",
        "z_csh_top", "z_sio2_bot", "z0", "deq",
        "z_csh_top_rel", "z_sio2_bot_rel", "z0_rel",
        "IF_WINDOW_HALF", "half_w_used", "frac_CSH_in_window",
        "anchor_CSH_z_aligned", "anchor_SiO2_z_aligned",
        "d_CSH_surface_to_anchor", "d_SiO2_surface_to_anchor",
        "n_Ca_in_window"
        ])
        for m in iface_metrics:
            w.writerow([
            int(m["frame"]),
            fnum(m["z_csh_top"]), fnum(m["z_sio2_bot"]), fnum(m["z0"]), fnum(m["deq"]),
            fnum(m["z_csh_top_rel"]), fnum(m["z_sio2_bot_rel"]), fnum(m["z0_rel"]),
            fnum(m["IF_WINDOW_HALF"]), fnum(m["half_w_used"]), fnum(m["frac_CSH_in_window"]),
            fnum(m.get("anchor_CSH_z_aligned", "")),
            fnum(m.get("anchor_SiO2_z_aligned", "")),
            fnum(m.get("d_CSH_surface_to_anchor", "")),
            fnum(m.get("d_SiO2_surface_to_anchor", "")),
            fnum(m["n_Ca_in_window"]),
            ])
        
    

    # 6) 配位数 CN（每帧 + 块平均）
    cn_keys = list(cn_per_frame[0].keys()) if cn_per_frame else []
    with open(os.path.join(out_dir, "cn_counts_per_frame.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(["frame"] + cn_keys)
        for i, fr in enumerate(frames[:len(cn_per_frame)]):
            w.writerow([int(fr)] + [fnum(cn_per_frame[i][k]) for k in cn_keys])

    with open(os.path.join(out_dir, "cn_counts_block_mean_se.csv"), "w", newline='', encoding="utf-8-sig") as f:
        w = csv.writer(f, lineterminator="\r\n")
        w.writerow(["pair", "mean", "se", "n_blocks"])
        n_blocks_cn = len(cn_blocks)
        for k in cn_keys:
            w.writerow([k, fnum(cn_mean.get(k, 0.0)), fnum(cn_se.get(k, 0.0)), int(n_blocks_cn)])

    # 7) 写 Excel 工作簿
    import pandas as pd
    excel_path = os.path.join(out_dir, "analysis_all.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # 7.1) 化学键
        if bonds_per_frame:
            df_bonds_frame = pd.DataFrame(bonds_per_frame)
            df_bonds_frame.insert(0, "frame", frames[:len(bonds_per_frame)])
            df_bonds_frame.to_excel(writer, sheet_name="bonds_per_frame", index=False)

            df_bonds_block = pd.DataFrame({
                "pair": bond_keys,
                "mean": [float(bond_mean.get(k, 0.0)) for k in bond_keys],
                "se":   [float(bond_se.get(k, 0.0))   for k in bond_keys],
                "n_blocks": [int(n_blocks_bonds)] * len(bond_keys)
            })
            df_bonds_block.to_excel(writer, sheet_name="bonds_block_mean", index=False)

        # 7.2) 氢键
        if hbonds_per_frame:
            df_hbonds_frame = pd.DataFrame(hbonds_per_frame)
            df_hbonds_frame.insert(0, "frame", frames[:len(hbonds_per_frame)])
            df_hbonds_frame.to_excel(writer, sheet_name="hbonds_per_frame", index=False)

            df_hbonds_block = pd.DataFrame({
                "pair": hbond_keys,
                "mean": [float(hbond_mean.get(k, 0.0)) for k in hbond_keys],
                "se":   [float(hbond_se.get(k, 0.0))   for k in hbond_keys],
                "n_blocks": [int(n_blocks_hb)] * len(hbond_keys)
            })
            df_hbonds_block.to_excel(writer, sheet_name="hbonds_block_mean", index=False)

        # 7.3) O–H 共价键
        if oh_bonds_per_frame:
            df_oh_frame = pd.DataFrame(oh_bonds_per_frame)
            df_oh_frame.insert(0, "frame", frames[:len(oh_bonds_per_frame)])
            df_oh_frame.to_excel(writer, sheet_name="oh_per_frame", index=False)

            df_oh_block = pd.DataFrame({
                "pair": oh_keys,
                "mean": [float(oh_mean.get(k, 0.0)) for k in oh_keys],
                "se":   [float(oh_se.get(k, 0.0))   for k in oh_keys],
                "n_blocks": [int(n_blocks_oh)] * len(oh_keys)
            })
            df_oh_block.to_excel(writer, sheet_name="oh_block_mean", index=False)

        # 7.4) RDF 合并
        if rdf_results:
            first_pair = next(iter(rdf_results.keys()))
            base_r = np.asarray(rdf_results[first_pair]["r"], dtype=float)
            df_rdf = pd.DataFrame({"r(Angstrom)": base_r})
            for pair_name, res in rdf_results.items():
                df_rdf[f"{pair_name}_g"]  = np.asarray(res["g_mean"], dtype=float)
                df_rdf[f"{pair_name}_se"] = np.asarray(res["g_se"], dtype=float)
            df_rdf.to_excel(writer, sheet_name="rdf_combined", index=False)

        # 7.5) CN（配位数）
        if cn_per_frame:
            df_cn_frame = pd.DataFrame(cn_per_frame)
            df_cn_frame.insert(0, "frame", frames[:len(cn_per_frame)])
            df_cn_frame.to_excel(writer, sheet_name="cn_per_frame", index=False)

            df_cn_block = pd.DataFrame({
                "pair": cn_keys,
                "mean": [float(cn_mean.get(k, 0.0)) for k in cn_keys],
                "se":   [float(cn_se.get(k, 0.0))   for k in cn_keys],
                "n_blocks": [int(len(cn_blocks))] * len(cn_keys)
            })
            df_cn_block.to_excel(writer, sheet_name="cn_block_mean", index=False)

        # 7.6) z 向密度（合并表）
        if dens_results:
            any_t = next(iter(dens_results.keys()))
            base_z = np.asarray(dens_results[any_t]["z"], dtype=float)
            df_dens = pd.DataFrame({"z(Angstrom)": base_z})
            for t in sorted(dens_results.keys(), key=lambda x: TYPE_NAME.get(x, f"type{x}")):
                name = TYPE_NAME.get(t, f"type{t}")
                df_dens[f"{name}_mean"] = np.asarray(dens_results[t]["rho_mean"], dtype=float)
                df_dens[f"{name}_se"]   = np.asarray(dens_results[t]["rho_se"],   dtype=float)
            df_dens.to_excel(writer, sheet_name="density_z_combined", index=False)

    print(f"Excel 工作簿已生成: {excel_path}")

    # ---------------- 简要总结 ----------------
    print("\n=== 分析完成 ===")
    print("化学键（块均值±SE）：")
    for k in bond_keys:
        print(f"  {k:30s}: {bond_mean.get(k,0):8.3f} ± {bond_se.get(k,0):6.3f}")
    print("\n氢键（块均值±SE）：")
    for k in hbond_keys:
        print(f"  {k:30s}: {hbond_mean.get(k,0):8.3f} ± {hbond_se.get(k,0):6.3f}")
    print(f"\nRDF 导出: {len(rdf_results)} 对，文件在 {out_dir}")
    print(f"ITZ 密度剖面导出: {len(dens_results)} 个类型，文件在 {out_dir}")
    print(f"帧数: {len(frames)}，块大小: {BLOCK_SIZE}")
    vals = [m["frac_CSH_in_window"] for m in iface_metrics]
    print(f"CSH_Si 覆盖率: mean={np.mean(vals):.4f}, se={np.std(vals, ddof=1)/math.sqrt(len(vals)) if len(vals)>1 else 0.0:.4f}")

if __name__ == "__main__":
    main()